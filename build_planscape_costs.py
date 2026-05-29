#!/usr/bin/env python3
"""Generate Planscape_Cost_Model.xlsx — an editable subscription/cost model."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ---- Palette ----
NAVY = "1A237E"
ORANGE = "E8912D"
GREEN = "2E7D32"
GREY = "ECEFF1"
LIGHT = "F5F5F5"
RED = "C62828"

hdr = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor=NAVY)
sub_fill = PatternFill("solid", fgColor=ORANGE)
sub_font = Font(bold=True, color="FFFFFF", size=11)
input_fill = PatternFill("solid", fgColor="FFF8E1")
calc_fill = PatternFill("solid", fgColor=LIGHT)
total_fill = PatternFill("solid", fgColor=GREEN)
total_font = Font(bold=True, color="FFFFFF", size=12)
bold = Font(bold=True)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="B0BEC5")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
link_font = Font(color="1565C0", underline="single")

def style_header_row(ws, row, ncols, fill=hdr_fill, font=hdr):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border

# =====================================================================
# SHEET 1: Inputs (the levers you change)
# =====================================================================
ws = wb.active
ws.title = "Inputs"
ws.sheet_properties.tabColor = ORANGE

ws["A1"] = "PLANSCAPE — COST MODEL INPUTS"
ws["A1"].font = Font(bold=True, size=16, color=NAVY)
ws.merge_cells("A1:C1")
ws["A2"] = "Edit the yellow cells. Everything on the 'Cost Model' tab recalculates automatically."
ws["A2"].font = Font(italic=True, size=10, color="555555")
ws.merge_cells("A2:C2")

ws["A4"] = "Parameter"; ws["B4"] = "Value"; ws["C4"] = "Notes"
style_header_row(ws, 4, 3)

inputs = [
    ("firms",            "Number of firms / tenants",         1,      "Tenants on the platform"),
    ("users",            "Total active users",                25,     "Across all firms"),
    ("projects",         "Total active projects",             10,     "Drives storage + AI usage"),
    ("storage_gb",       "Object storage used (GB)",          50,     "Models, attachments, thumbnails"),
    ("egress_gb",        "Monthly data egress (GB)",          20,     "Downloads / viewer streaming"),
    ("emails_month",     "Transactional emails / month",      8000,   "Invites, resets, dunning, reminders"),
    ("ai_tokens_m",      "Azure OpenAI tokens / month (millions)", 5, "Auto-linking + classification"),
    ("ocr_pages",        "OCR pages / month",                 4000,   "Document text extraction"),
    ("map_loads",        "Mapbox map loads / month",          10000,  "Marketing-site / viewer map"),
    ("revenue_month",    "Monthly revenue processed ($)",     20000,  "Subscriptions billed to customers"),
    ("intl_share",       "Share of revenue that is international (%)", 30, "Higher card fees"),
    ("sso_users",        "Users needing enterprise SSO",      0,      "Okta/Entra per-user add-on"),
    ("model_conversions","3D model conversions / month (APS)",0,      "0 = use free IfcOpenShell instead"),
]
r = 5
named = {}
for key, label, val, note in inputs:
    ws.cell(row=r, column=1, value=label).font = bold
    c = ws.cell(row=r, column=2, value=val)
    c.fill = input_fill; c.border = border; c.alignment = center
    ws.cell(row=r, column=3, value=note).alignment = wrap
    ws.cell(row=r, column=1).border = border
    ws.cell(row=r, column=3).border = border
    # define a workbook-level named range pointing at this cell
    ref = f"Inputs!${'B'}${r}"
    wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(key, attr_text=ref))
    named[key] = ref
    r += 1

ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 42

# Assumptions block
ar = r + 1
ws.cell(row=ar, column=1, value="UNIT-RATE ASSUMPTIONS (edit to match your contracts)").font = Font(bold=True, color=NAVY, size=12)
ws.merge_cells(start_row=ar, start_column=1, end_row=ar, end_column=3)
ar += 1
ws.cell(row=ar, column=1, value="Rate"); ws.cell(row=ar, column=2, value="Value"); ws.cell(row=ar, column=3, value="Unit")
style_header_row(ws, ar, 3)
ar += 1
rates = [
    ("rate_stripe_dom",  "Stripe domestic fee", 0.029, "fraction of revenue"),
    ("rate_stripe_intl", "Stripe international fee", 0.0325, "fraction of revenue"),
    ("rate_s3_gb",       "Storage $/GB-month", 0.015, "R2 ~0.015 / S3 ~0.023"),
    ("rate_egress_gb",   "Egress $/GB", 0.05, "R2=0 / S3~0.09 / DO included"),
    ("rate_email_k",     "Email $ per 1,000", 1.25, "Postmark tiered"),
    ("rate_ai_m",        "Azure OpenAI $ per 1M tokens (blended)", 4.0, "GPT-4o blended in/out"),
    ("rate_ocr_k",       "OCR $ per 1,000 pages", 1.25, "Azure AI Vision S1"),
    ("rate_map_k",       "Mapbox $ per 1,000 loads over free tier", 0.50, "free 50k/mo"),
    ("rate_sso_user",    "SSO $/user-month", 3.0, "Okta ~2 / Entra P1 ~6"),
    ("rate_aps_conv",    "APS $/model conversion (flex tokens)", 1.5, "approx"),
    ("rate_hosting",     "Hosting base $/month", 120.0, "VPS + managed PG/Redis"),
    ("rate_postmark_base","Email platform base $/month", 15.0, "Postmark Basic"),
    ("rate_eas",         "Expo EAS $/month", 99.0, "0 if free tier"),
    ("rate_sentry",      "Sentry $/month", 26.0, "0 if free tier"),
]
for key, label, val, unit in rates:
    ws.cell(row=ar, column=1, value=label).font = bold
    c = ws.cell(row=ar, column=2, value=val); c.fill = input_fill; c.border = border; c.alignment = center
    ws.cell(row=ar, column=3, value=unit).alignment = wrap
    ws.cell(row=ar, column=1).border = border; ws.cell(row=ar, column=3).border = border
    wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(key, attr_text=f"Inputs!$B${ar}"))
    ar += 1

# =====================================================================
# SHEET 2: Cost Model (live calc)
# =====================================================================
cm = wb.create_sheet("Cost Model")
cm.sheet_properties.tabColor = GREEN
cm["A1"] = "PLANSCAPE — MONTHLY COST MODEL"
cm["A1"].font = Font(bold=True, size=16, color=NAVY)
cm.merge_cells("A1:F1")
cm["A2"] = "Costs recalc from the Inputs tab. Yellow = pulled from inputs; grey = computed."
cm["A2"].font = Font(italic=True, size=10, color="555555")
cm.merge_cells("A2:F2")

cols = ["Service", "Category", "Cost driver", "Monthly cost ($)", "Annual ($)", "Pricing link"]
for i, h in enumerate(cols, 1):
    cm.cell(row=4, column=i, value=h)
style_header_row(cm, 4, 6)

# Each row: (service, category, driver text, monthly formula, link)
rows = [
    ("Hosting (VPS + managed PG/Redis)", "Required", "Flat base", "=rate_hosting", "https://www.digitalocean.com/pricing"),
    ("Object storage", "Required", "Per GB stored", "=storage_gb*rate_s3_gb", "https://developers.cloudflare.com/r2/pricing/"),
    ("Data egress", "Required", "Per GB out", "=egress_gb*rate_egress_gb", "https://aws.amazon.com/s3/pricing/"),
    ("Stripe (domestic)", "Required", "% of domestic revenue", "=revenue_month*(1-intl_share/100)*rate_stripe_dom", "https://stripe.com/pricing"),
    ("Stripe (international)", "Required", "% of intl revenue", "=revenue_month*(intl_share/100)*rate_stripe_intl", "https://stripe.com/pricing"),
    ("Email platform base", "Required", "Flat base", "=rate_postmark_base", "https://postmarkapp.com/pricing"),
    ("Email volume", "Required", "Per 1,000 emails", "=MAX(0,emails_month-10000)/1000*rate_email_k", "https://postmarkapp.com/pricing"),
    ("Azure OpenAI", "AI", "Per 1M tokens", "=ai_tokens_m*rate_ai_m", "https://azure.microsoft.com/en-us/pricing/details/cognitive-services/openai-service/"),
    ("Azure AI Vision (OCR)", "AI", "Per 1,000 pages", "=MAX(0,ocr_pages-5000)/1000*rate_ocr_k", "https://azure.microsoft.com/en-us/pricing/details/cognitive-services/computer-vision/"),
    ("Firebase Cloud Messaging", "Push", "Flat (free)", "=0", "https://firebase.google.com/pricing"),
    ("Mapbox", "Maps", "Per 1,000 loads over free", "=MAX(0,map_loads-50000)/1000*rate_map_k", "https://www.mapbox.com/pricing"),
    ("Autodesk APS (3D conversion)", "Optional", "Per conversion", "=model_conversions*rate_aps_conv", "https://aps.autodesk.com/pricing"),
    ("Enterprise SSO (Okta/Entra)", "Optional", "Per SSO user", "=sso_users*rate_sso_user", "https://www.okta.com/pricing/"),
    ("Expo EAS (mobile builds)", "Mobile", "Flat", "=rate_eas", "https://expo.dev/pricing"),
    ("Sentry (crash reporting)", "Optional", "Flat", "=rate_sentry", "https://sentry.io/pricing/"),
]
start = 5
rr = start
for svc, cat, driver, formula, link in rows:
    cm.cell(row=rr, column=1, value=svc).font = bold
    cm.cell(row=rr, column=2, value=cat)
    cm.cell(row=rr, column=3, value=driver).alignment = wrap
    mc = cm.cell(row=rr, column=4, value=formula)
    mc.number_format = '#,##0.00'; mc.fill = calc_fill
    ac = cm.cell(row=rr, column=5, value=f"=D{rr}*12")
    ac.number_format = '#,##0.00'; ac.fill = calc_fill
    lk = cm.cell(row=rr, column=6, value=link); lk.hyperlink = link; lk.font = link_font
    for c in range(1, 7):
        cm.cell(row=rr, column=c).border = border
    if cat == "Optional":
        cm.cell(row=rr, column=2).font = Font(color=RED, italic=True)
    rr += 1

# Totals
tot = rr + 1
cm.cell(row=tot, column=1, value="MONTHLY TOTAL").font = total_font
cm.cell(row=tot, column=1).fill = total_fill
cm.cell(row=tot, column=2).fill = total_fill
cm.cell(row=tot, column=3).fill = total_fill
tm = cm.cell(row=tot, column=4, value=f"=SUM(D{start}:D{rr-1})")
tm.number_format = '#,##0.00'; tm.font = total_font; tm.fill = total_fill
ta = cm.cell(row=tot, column=5, value=f"=SUM(E{start}:E{rr-1})")
ta.number_format = '#,##0.00'; ta.font = total_font; ta.fill = total_fill
cm.cell(row=tot, column=6).fill = total_fill

# Per-firm / per-user / per-project derived
d = tot + 2
cm.cell(row=d, column=1, value="Cost per firm / month").font = bold
cm.cell(row=d, column=4, value=f"=D{tot}/MAX(1,firms)").number_format = '#,##0.00'
cm.cell(row=d+1, column=1, value="Cost per user / month").font = bold
cm.cell(row=d+1, column=4, value=f"=D{tot}/MAX(1,users)").number_format = '#,##0.00'
cm.cell(row=d+2, column=1, value="Cost per project / month").font = bold
cm.cell(row=d+2, column=4, value=f"=D{tot}/MAX(1,projects)").number_format = '#,##0.00'

cm.cell(row=d+4, column=1, value="One-off costs (not in monthly total):").font = Font(bold=True, color=NAVY)
cm.cell(row=d+5, column=1, value="Apple Developer Program")
cm.cell(row=d+5, column=4, value="$99 / year");
lk = cm.cell(row=d+5, column=6, value="https://developer.apple.com/programs/"); lk.hyperlink="https://developer.apple.com/programs/"; lk.font=link_font
cm.cell(row=d+6, column=1, value="Google Play Console")
cm.cell(row=d+6, column=4, value="$25 once")
lk = cm.cell(row=d+6, column=6, value="https://play.google.com/console/signup"); lk.hyperlink="https://play.google.com/console/signup"; lk.font=link_font

cm.column_dimensions["A"].width = 34
cm.column_dimensions["B"].width = 12
cm.column_dimensions["C"].width = 24
cm.column_dimensions["D"].width = 16
cm.column_dimensions["E"].width = 14
cm.column_dimensions["F"].width = 60

# =====================================================================
# SHEET 3: Paid Services (reference)
# =====================================================================
ps = wb.create_sheet("Paid Services")
ps.sheet_properties.tabColor = NAVY
ps["A1"] = "PAID SERVICES — REFERENCE (indicative, early 2026 — verify on link)"
ps["A1"].font = Font(bold=True, size=14, color=NAVY)
ps.merge_cells("A1:F1")
ph = ["Service", "Category", "Scaling driver", "Indicative price", "Pricing link", "Sign-up link"]
for i, h in enumerate(ph, 1):
    ps.cell(row=3, column=i, value=h)
style_header_row(ps, 3, 6)

paid = [
    ("Stripe","Payments (req)","% per transaction","2.9% + $0.30 US; 3.25%+20p intl","https://stripe.com/pricing","https://dashboard.stripe.com/register"),
    ("Flutterwave","Payments (req)","% per transaction","~1.4% local / ~3.8% intl","https://flutterwave.com/pricing","https://dashboard.flutterwave.com/signup"),
    ("Azure OpenAI","AI (req)","per 1M tokens","GPT-4o ~$2.50 in/$10 out; mini ~$0.15/$0.60","https://azure.microsoft.com/en-us/pricing/details/cognitive-services/openai-service/","https://azure.microsoft.com/free"),
    ("Azure AI Vision (OCR)","AI (req)","per 1,000 pages","Free 5k/mo; ~$1.00-1.50 per 1k","https://azure.microsoft.com/en-us/pricing/details/cognitive-services/computer-vision/","https://azure.microsoft.com/free"),
    ("Firebase FCM","Push (req)","flat","$0 free","https://firebase.google.com/pricing","https://console.firebase.google.com"),
    ("Postmark","Email (req)","per emails/mo","$15/mo for 10k, tiered","https://postmarkapp.com/pricing","https://account.postmarkapp.com/sign_up"),
    ("AWS S3","Storage (req)","per GB + egress","~$0.023/GB-mo + ~$0.09/GB egress","https://aws.amazon.com/s3/pricing/","https://aws.amazon.com"),
    ("Cloudflare R2","Storage (req-alt)","per GB, no egress","$0.015/GB-mo, $0 egress","https://developers.cloudflare.com/r2/pricing/","https://dash.cloudflare.com/sign-up"),
    ("DigitalOcean Spaces","Storage (req-alt)","flat + overage","$5/mo 250GB +1TB transfer","https://www.digitalocean.com/pricing/spaces-object-storage","https://cloud.digitalocean.com/registrations/new"),
    ("Autodesk APS / Forge","3D (opt)","flex tokens/conversion","Pay-as-you-go cloud credits","https://aps.autodesk.com/pricing","https://aps.autodesk.com"),
    ("Sentry","Monitoring (opt)","per events/users","Free dev; Team $26/mo","https://sentry.io/pricing","https://sentry.io/signup"),
    ("Mapbox","Maps","per map loads","Free 50k/mo then per 1k","https://www.mapbox.com/pricing","https://account.mapbox.com/auth/signup"),
    ("Expo EAS","Mobile builds","per build mins/seats","Free limited; Production $99/mo","https://expo.dev/pricing","https://expo.dev/signup"),
    ("Apple Developer","Mobile (req to ship iOS)","flat/year","$99/year","https://developer.apple.com/programs","https://developer.apple.com/programs/enroll"),
    ("Google Play Console","Mobile (req to ship Android)","one-time","$25 once","https://play.google.com/console/signup","https://play.google.com/console/signup"),
    ("Slack","Webhooks (opt)","per user/mo","Free tier; Pro ~$7.25/user","https://slack.com/pricing","https://slack.com/get-started"),
    ("Microsoft Teams","Webhooks (opt)","M365 seat","From ~$4/user (Business Basic)","https://www.microsoft.com/microsoft-teams","https://www.microsoft.com/microsoft-365"),
    ("Okta SSO","Identity (opt)","per user/mo","SSO ~$2/user/mo","https://www.okta.com/pricing/","https://www.okta.com/free-trial/"),
    ("Microsoft Entra ID P1","Identity (opt)","per user/mo","~$6/user/mo","https://www.microsoft.com/security/business/microsoft-entra-pricing","https://azure.microsoft.com/free"),
    ("Seq","Logs (opt)","per seat/license","Free single-user; Team paid","https://datalust.co/pricing","https://datalust.co"),
]
rr = 4
for row in paid:
    for i, v in enumerate(row, 1):
        cell = ps.cell(row=rr, column=i, value=v)
        cell.border = border
        cell.alignment = wrap
        if i == 1: cell.font = bold
        if i in (5, 6) and v.startswith("http"):
            cell.hyperlink = v; cell.font = link_font
    rr += 1
widths = [26, 22, 18, 32, 58, 50]
for i, w in enumerate(widths, 1):
    ps.column_dimensions[get_column_letter(i)].width = w

# =====================================================================
# SHEET 4: Free / Self-hosted (reference)
# =====================================================================
fs = wb.create_sheet("Free & Self-Hosted")
fs.sheet_properties.tabColor = "607D8B"
fs["A1"] = "FREE / OPEN-SOURCE / SELF-HOSTABLE — NO SUBSCRIPTION"
fs["A1"].font = Font(bold=True, size=14, color=GREEN)
fs.merge_cells("A1:C1")
for i, h in enumerate(["Service", "Role in Planscape", "Note"], 1):
    fs.cell(row=3, column=i, value=h)
style_header_row(fs, 3, 3, fill=PatternFill("solid", fgColor=GREEN))

free = [
    ("PostgreSQL 16","Primary database","Self-host; pay only if managed"),
    ("Redis 7","Cache, SignalR backplane, jobs","Self-host"),
    ("MinIO","S3-compatible storage","Free alternative to S3/R2"),
    ("Local filesystem","Storage fallback","Free"),
    ("Hangfire","Background jobs","Postgres-backed, free"),
    ("SignalR","Real-time hubs","Built into ASP.NET Core"),
    ("IfcOpenShell / ifcconvert","3D IFC->glTF conversion","Free alternative to APS"),
    ("Expo Go","Dev runtime","Free (only EAS builds cost)"),
    ("Expo Push","Push notifications","Free"),
    ("Generic SMTP","Email","Free alternative to Postmark"),
    ("Prometheus","Metrics","Self-host, opt-in"),
    ("Grafana","Dashboards","Self-host, opt-in"),
    ("OpenTelemetry Collector","Tracing","Self-host, opt-in"),
    ("Elasticsearch / OpenSearch","Log search","Self-host, opt-in"),
    ("ClamAV","Antivirus on upload","Self-host, opt-in (off by default)"),
    ("QuestPDF","Invoice PDF generation","Free for revenue < $1M"),
    ("ImageSharp / SkiaSharp","Image processing, redaction","Open-source NuGet"),
    ("ONNX Runtime","Face/plate detection","Open-source"),
    ("xbim.Essentials","IFC parsing","Open-source"),
    ("Next.js","Web dashboard","Self-hostable"),
]
rr = 4
for row in free:
    for i, v in enumerate(row, 1):
        cell = fs.cell(row=rr, column=i, value=v)
        cell.border = border; cell.alignment = wrap
        if i == 1: cell.font = bold
    rr += 1
for i, w in enumerate([28, 32, 44], 1):
    fs.column_dimensions[get_column_letter(i)].width = w

# =====================================================================
# SHEET 5: Scenarios (snapshot reference)
# =====================================================================
sc = wb.create_sheet("Scenarios")
sc.sheet_properties.tabColor = "8E24AA"
sc["A1"] = "INDICATIVE MONTHLY COST BY SCALE"
sc["A1"].font = Font(bold=True, size=14, color=NAVY)
sc.merge_cells("A1:D1")
sc["A2"] = "Plug these into the Inputs tab to get exact numbers."
sc["A2"].font = Font(italic=True, size=10, color="555555")
sc.merge_cells("A2:D2")
for i, h in enumerate(["Scale", "Profile", "Main paid items", "Indicative monthly"], 1):
    sc.cell(row=4, column=i, value=h)
style_header_row(sc, 4, 4)
scen = [
    ("Solo / pilot","1 firm, <=5 users, 1 project","VPS, free FCM/Expo, free-tier Azure, Mapbox free; Stripe % only","$30-80 + ~3% revenue + $99/yr Apple"),
    ("Small","<=25 users, ~10 projects","Managed PG/Redis, R2/Spaces, Postmark $15, modest Azure, EAS $99","$200-350 + payment %"),
    ("Mid","Multi-firm, 100+ users","Bigger DB, more storage, higher Azure spend, Sentry, Slack Pro","$600-1,200 + payment %"),
    ("Enterprise","Unlimited, SSO required","HA PG/Redis, large storage/egress, heavy Azure AI, Okta/Entra SSO, APS, Seq/Grafana","$3k-10k+ + payment % + per-user SSO"),
]
rr = 5
for row in scen:
    for i, v in enumerate(row, 1):
        cell = sc.cell(row=rr, column=i, value=v)
        cell.border = border; cell.alignment = wrap
        if i == 1: cell.font = bold
    rr += 1
for i, w in enumerate([16, 30, 46, 34], 1):
    sc.column_dimensions[get_column_letter(i)].width = w

out = "/home/user/STINGTOOLS/Planscape_Cost_Model.xlsx"
wb.save(out)
print("Saved", out)
print("Sheets:", wb.sheetnames)
