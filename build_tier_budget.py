#!/usr/bin/env python3
"""Generate Planscape_Tier_Budget.xlsx — multi-tier budget + growth guidance."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

NAVY="1A237E"; ORANGE="E8912D"; GREEN="2E7D32"; RED="C62828"; LIGHT="F5F5F5"; YEL="FFF8E1"
hdr=Font(bold=True,color="FFFFFF",size=11); hdr_fill=PatternFill("solid",fgColor=NAVY)
bold=Font(bold=True)
wrap=Alignment(wrap_text=True,vertical="top"); ctr=Alignment(horizontal="center",vertical="center")
thin=Side(style="thin",color="B0BEC5"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
input_fill=PatternFill("solid",fgColor=YEL); calc_fill=PatternFill("solid",fgColor=LIGHT)
total_fill=PatternFill("solid",fgColor=GREEN); total_font=Font(bold=True,color="FFFFFF",size=11)
link_font=Font(color="1565C0",underline="single")
money='#,##0.00'

wb=openpyxl.Workbook()

def hrow(ws,row,n,fill=hdr_fill,font=hdr):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c); cell.fill=fill; cell.font=font
        cell.alignment=Alignment(wrap_text=True,vertical="center"); cell.border=border

TIERS=["Pilot","Starter","Growth","Scale","Enterprise"]
COUNTS={
 "Firms":[2,5,20,50,150],
 "Projects":[10,50,400,1500,5000],
 "Users (coordinators)":[20,100,1000,5000,15000],
 "Object storage billed (GB)":[0,0,1000,5000,20000],
 "SSO users":[0,0,0,0,15000],
}
# (label, note, type, values) — type: fixed|storage|sso
COSTS=[
 ("App / server compute","VPS or droplets (API+worker)","fixed",[15,48,96,400,1500]),
 ("Database (PostgreSQL)","In-box early; managed + standby later","fixed",[0,0,120,480,2000]),
 ("Cache (Redis)","In-box early; managed later","fixed",[0,0,15,60,300]),
 ("Object storage","Local disk(0) early; R2/Spaces later","storage",None),
 ("Connection pooling (PgBouncer)","Self-hosted in compose","fixed",[0,0,0,0,0]),
 ("Email sending","Free tier early; Postmark/SES later","fixed",[0,15,15,50,150]),
 ("Push (Firebase + Expo)","Free at every tier","fixed",[0,0,0,0,0]),
 ("Domain + TLS","Domain ~$12/yr; Let's Encrypt free","fixed",[1,1,2,2,5]),
 ("Backups","Snapshots / managed backup","fixed",[3,10,28,120,400]),
 ("Payments (Stripe/Flutterwave)","% per txn, no monthly fee","fixed",[0,0,0,0,0]),
 ("Azure OpenAI (smart-linking)","Usage-based; optional feature","fixed",[0,0,40,400,2000]),
 ("Azure AI Vision (OCR)","Usage-based; optional feature","fixed",[0,0,10,80,400]),
 ("Crash reporting (Sentry)","Free tier early","fixed",[0,0,0,26,80]),
 ("Observability (Seq/Grafana)","Self-host free; managed at top","fixed",[0,0,0,0,200]),
 ("Enterprise SSO (Okta/Entra)","Per-user; enterprise only","sso",None),
 ("3D conversion (APS/Forge)","Optional; IfcOpenShell is free","fixed",[0,0,0,0,0]),
 ("Mobile app stores (Apple/Google/EAS)","Web-first = $0; add for native apps","fixed",[0,0,0,0,0]),
]

# ============ SHEET 1: How to Use ============
ws=wb.active; ws.title="How to Use"; ws.sheet_properties.tabColor=ORANGE
ws["A1"]="PLANSCAPE — TIER BUDGET & GROWTH PLAN"; ws["A1"].font=Font(bold=True,size=16,color=NAVY)
notes=[
 "",
 "HOW TO USE THIS WORKBOOK",
 "  - 'Tier Budget': monthly + annual cost per growth tier. Edit YELLOW cells (counts + rates).",
 "  - 'When to Add What': the trigger to move to the next tier and exactly what to provision.",
 "  - 'Services Catalog': every service - free vs paid, when it enters, real price, pricing + signup links.",
 "  - 'Deploy by Tier': the architecture at each stage (one box -> managed -> HA -> multi-node).",
 "",
 "GUIDING PRINCIPLES",
 "  1. Start web-first. Skip Apple ($99/yr) + Google Play ($25) until native apps - send links / PWA.",
 "  2. One box until it hurts. Run the whole docker-compose on a single VPS through the Starter tier.",
 "  3. Split the database first. PostgreSQL is the first thing to outgrow the box - move to managed at Growth.",
 "  4. Storage is the only cost that climbs with projects. Local disk until ~50 GB, then R2 (zero egress).",
 "  5. AI is opt-in. Azure OpenAI/Vision only cost money when you switch on smart-linking / OCR.",
 "  6. Payments add no monthly fee - only a % per transaction. Turn on when you bill in-app.",
 "  7. Add HA + observability + SSO only at Scale/Enterprise, when downtime and audits cost real money.",
 "",
 "Prices are indicative (early 2026). Confirm on the linked pricing pages and overwrite the yellow rate cells.",
]
r=2
for line in notes:
    c=ws.cell(row=r,column=1,value=line)
    if line.strip().isupper() and line.strip(): c.font=Font(bold=True,color=NAVY,size=12)
    elif line.strip(): c.font=Font(size=11)
    r+=1
ws.column_dimensions["A"].width=112

# ============ SHEET 2: Tier Budget ============
tb=wb.create_sheet("Tier Budget"); tb.sheet_properties.tabColor=GREEN
tb["A1"]="MONTHLY BUDGET BY TIER (edit yellow cells)"; tb["A1"].font=Font(bold=True,size=14,color=NAVY)
tb["A2"]="Storage $/GB-mo:"; tb["B2"]=0.015; tb["A3"]="SSO $/user-mo:"; tb["B3"]=3.0
tb["A2"].font=bold; tb["A3"].font=bold
for cell in ("B2","B3"): tb[cell].fill=input_fill; tb[cell].border=border; tb[cell].alignment=ctr
wb.defined_names.add(DefinedName("rate_storage_gb",attr_text="'Tier Budget'!$B$2"))
wb.defined_names.add(DefinedName("rate_sso_user",attr_text="'Tier Budget'!$B$3"))

head=5
tb.cell(row=head,column=1,value="Item"); tb.cell(row=head,column=2,value="Note / driver")
for i,t in enumerate(TIERS): tb.cell(row=head,column=3+i,value=t)
hrow(tb,head,7)

row=head+1
count_rows={}
tb.cell(row=row,column=1,value="SCALE (editable)").font=Font(bold=True,color=ORANGE)
for c in range(1,8): tb.cell(row=row,column=c).fill=PatternFill("solid",fgColor="FFF3E0")
row+=1
for label,vals in COUNTS.items():
    tb.cell(row=row,column=1,value=label).font=bold; tb.cell(row=row,column=1).border=border
    tb.cell(row=row,column=2).border=border
    for i,v in enumerate(vals):
        cell=tb.cell(row=row,column=3+i,value=v); cell.fill=input_fill; cell.border=border; cell.alignment=ctr
    count_rows[label]=row; row+=1
storr=count_rows["Object storage billed (GB)"]; ssor=count_rows["SSO users"]

tb.cell(row=row,column=1,value="MONTHLY COSTS ($)").font=Font(bold=True,color=NAVY)
for c in range(1,8): tb.cell(row=row,column=c).fill=PatternFill("solid",fgColor="E8EAF6")
row+=1
cost_start=row
for label,note,typ,vals in COSTS:
    tb.cell(row=row,column=1,value=label).font=bold; tb.cell(row=row,column=1).border=border
    nc=tb.cell(row=row,column=2,value=note); nc.alignment=wrap; nc.border=border
    if note.lower().startswith("optional") or "optional" in note.lower():
        nc.font=Font(italic=True,color=RED,size=9)
    for i in range(5):
        col=3+i; L=get_column_letter(col)
        if typ=="fixed":
            cell=tb.cell(row=row,column=col,value=vals[i]); cell.fill=input_fill
        elif typ=="storage":
            cell=tb.cell(row=row,column=col,value=f"={L}{storr}*rate_storage_gb"); cell.fill=calc_fill
        elif typ=="sso":
            cell=tb.cell(row=row,column=col,value=f"={L}{ssor}*rate_sso_user"); cell.fill=calc_fill
        cell.number_format=money; cell.border=border; cell.alignment=ctr
    row+=1
cost_end=row-1

# totals
tb.cell(row=row,column=1,value="MONTHLY TOTAL").font=total_font
for i in range(5):
    L=get_column_letter(3+i)
    c=tb.cell(row=row,column=3+i,value=f"=SUM({L}{cost_start}:{L}{cost_end})")
    c.number_format=money; c.font=total_font; c.fill=total_fill; c.border=border
tb.cell(row=row,column=1).fill=total_fill; tb.cell(row=row,column=2).fill=total_fill; tb.cell(row=row,column=2).border=border
tb.cell(row=row,column=1).border=border
total_row=row; row+=1
tb.cell(row=row,column=1,value="ANNUAL TOTAL").font=bold; tb.cell(row=row,column=1).border=border; tb.cell(row=row,column=2).border=border
for i in range(5):
    L=get_column_letter(3+i)
    c=tb.cell(row=row,column=3+i,value=f"={L}{total_row}*12"); c.number_format=money; c.font=bold; c.border=border; c.alignment=ctr
annual_row=row; row+=2

# per-unit
fr=count_rows["Firms"]; pr=count_rows["Projects"]; ur=count_rows["Users (coordinators)"]
for label,den in [("Cost per firm / month",fr),("Cost per user / month",ur),("Cost per project / month",pr)]:
    tb.cell(row=row,column=1,value=label).font=bold; tb.cell(row=row,column=1).border=border; tb.cell(row=row,column=2).border=border
    for i in range(5):
        L=get_column_letter(3+i)
        c=tb.cell(row=row,column=3+i,value=f"={L}{total_row}/MAX(1,{L}{den})"); c.number_format=money; c.border=border; c.alignment=ctr
    row+=1
row+=1
tb.cell(row=row,column=1,value="Plus: payments = % per transaction (Stripe 2.9%+$0.30 domestic / 3.25%+20p intl; Flutterwave ~1.4% local).").font=Font(italic=True,size=9,color="555555")
tb.merge_cells(start_row=row,start_column=1,end_row=row,end_column=7)
row+=1
tb.cell(row=row,column=1,value="One-off when you go native: Apple $99/yr + Google Play $25 once. Web-first launch needs neither.").font=Font(italic=True,size=9,color="555555")
tb.merge_cells(start_row=row,start_column=1,end_row=row,end_column=7)

tb.column_dimensions["A"].width=34; tb.column_dimensions["B"].width=34
for i in range(5): tb.column_dimensions[get_column_letter(3+i)].width=13
tb.freeze_panes="C6"

# ============ SHEET 3: When to Add What ============
wa=wb.create_sheet("When to Add What"); wa.sheet_properties.tabColor="6A1B9A"
wa["A1"]="GROWTH ROADMAP — TRIGGERS & ACTIONS"; wa["A1"].font=Font(bold=True,size=14,color=NAVY)
wa.merge_cells("A1:E1")
for i,h in enumerate(["Tier","Typical scale","Trigger to enter","What to add / change","Approx monthly"],1):
    wa.cell(row=3,column=i,value=h)
hrow(wa,3,5)
road=[
 ("Pilot","2 firms / 10 proj / 20 users","Day one. Validating with friendly firms.",
  "1 small VPS running full docker-compose; local-disk storage; free email tier (Brevo/SES); free push; web links only.","~$16"),
 ("Starter","5 firms / 50 proj / 100 users","First paying customers; >~20 GB of files.",
  "Bigger VPS (4 vCPU/16 GB); move files to R2/Spaces; Postmark for deliverability; turn on Stripe/Flutterwave billing.","~$60-90"),
 ("Growth","20 firms / 400 proj / 1,000 users","Single box strains; need uptime; ~50-200 concurrent.",
  "Move PostgreSQL to managed; managed Redis; separate worker node; nightly backups; enable Azure AI features if selling them.","~$280-400"),
 ("Scale","50 firms / 1,500 proj / 5,000 users","Downtime now costs money; multi-region users.",
  "HA Postgres (primary+standby); 2x app nodes + load balancer; PgBouncer pooling; Sentry; observability stack; 5 TB storage.","~$1,200-2,500"),
 ("Enterprise","150+ firms / 5,000+ proj / 15,000+ users","Procurement, audits, SSO mandates, SLAs.",
  "SSO (Okta/Entra, per-user); SAML/SCIM; APS high-fidelity 3D; managed logs/SIEM; ClamAV; dedicated DB tier; 24/7 on-call.","$5k-12k+"),
]
r=4
for tier,scale,trig,act,cost in road:
    wa.cell(row=r,column=1,value=tier).font=bold
    wa.cell(row=r,column=2,value=scale); wa.cell(row=r,column=3,value=trig)
    wa.cell(row=r,column=4,value=act); wa.cell(row=r,column=5,value=cost).font=bold
    for c in range(1,6): wa.cell(row=r,column=c).border=border; wa.cell(row=r,column=c).alignment=wrap
    r+=1
for col,w in zip("ABCDE",[14,26,34,52,16]): wa.column_dimensions[col].width=w

# checklist of "what to add when feature X needed"
r+=1
wa.cell(row=r,column=1,value="FEATURE-DRIVEN ADD-ONS (independent of tier)").font=Font(bold=True,color=NAVY,size=12)
wa.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); r+=1
for i,h in enumerate(["When you want to...","Add this","Cost","",""],1): wa.cell(row=r,column=i,value=h)
hrow(wa,r,5); r+=1
addons=[
 ("Bill customers inside the app","Stripe (global) + Flutterwave (Africa)","% per txn, $0/mo"),
 ("Ship a native iOS/Android app","Apple Developer + Google Play + Expo EAS","$99/yr + $25 + ~$99/mo"),
 ("Auto-link elements / NLP / OCR","Azure OpenAI + Azure AI Vision","usage-based"),
 ("High-fidelity 3D model viewer","Autodesk APS (else free IfcOpenShell)","flex tokens"),
 ("Enterprise customer SSO","Okta or Microsoft Entra ID P1","$2-6/user/mo"),
 ("Crash visibility on mobile","Sentry","free tier / $26/mo"),
 ("Push to Slack/Teams channels","Webhook URLs (free)","$0"),
 ("Structured log search + alerts","Seq or Elastic/OpenSearch (self-host free)","$0 self-host"),
]
for w_,add,cost in addons:
    wa.cell(row=r,column=1,value=w_); wa.cell(row=r,column=2,value=add).font=bold; wa.cell(row=r,column=3,value=cost)
    for c in range(1,4): wa.cell(row=r,column=c).border=border; wa.cell(row=r,column=c).alignment=wrap
    r+=1

# ============ SHEET 4: Services Catalog ============
sc=wb.create_sheet("Services Catalog"); sc.sheet_properties.tabColor=NAVY
sc["A1"]="SERVICES CATALOG (indicative early-2026 prices)"; sc["A1"].font=Font(bold=True,size=14,color=NAVY)
sc.merge_cells("A1:G1")
for i,h in enumerate(["Service","Free or Paid","Enters at tier","Scaling driver","Indicative price","Pricing link","Sign-up link"],1):
    sc.cell(row=3,column=i,value=h)
hrow(sc,3,7)
cat=[
 ("VPS / compute (Hetzner/DO)","Paid","Pilot","per server size","$15-1,500/mo","https://www.hetzner.com/cloud","https://accounts.hetzner.com/signUp"),
 ("PostgreSQL","Free self-host / Paid managed","Growth (managed)","CPU/RAM/disk","$0 -> $120-2,000/mo","https://www.digitalocean.com/pricing/managed-databases","https://cloud.digitalocean.com/registrations/new"),
 ("Redis / Valkey","Free self-host / Paid managed","Growth (managed)","memory","$0 -> $15-300/mo","https://www.digitalocean.com/pricing/managed-databases","https://cloud.digitalocean.com/registrations/new"),
 ("Object storage (Cloudflare R2)","Paid","Growth","per GB, no egress","$0.015/GB-mo","https://developers.cloudflare.com/r2/pricing/","https://dash.cloudflare.com/sign-up"),
 ("Object storage (AWS S3)","Paid","Growth","per GB + egress","~$0.023/GB + egress","https://aws.amazon.com/s3/pricing/","https://aws.amazon.com"),
 ("Object storage (DO Spaces)","Paid","Growth","flat + overage","$5/mo 250GB","https://www.digitalocean.com/pricing/spaces-object-storage","https://cloud.digitalocean.com/registrations/new"),
 ("MinIO (self-host storage)","Free","Pilot","disk only","$0","https://min.io/pricing","https://min.io"),
 ("Email - Brevo / AWS SES","Free tier / Paid","Pilot","per email","free 300/day / $0.10 per 1k","https://www.brevo.com/pricing/","https://www.brevo.com"),
 ("Email - Postmark","Paid","Starter","per emails/mo","$15/mo for 10k","https://postmarkapp.com/pricing","https://account.postmarkapp.com/sign_up"),
 ("Firebase FCM (push)","Free","Pilot","flat","$0","https://firebase.google.com/pricing","https://console.firebase.google.com"),
 ("Expo Push","Free","Pilot","flat","$0","https://docs.expo.dev/push-notifications/overview/","https://expo.dev/signup"),
 ("Stripe","Paid","Starter (billing)","% per txn","2.9%+$0.30 / 3.25%+20p intl","https://stripe.com/pricing","https://dashboard.stripe.com/register"),
 ("Flutterwave","Paid","Starter (billing)","% per txn","~1.4% local / ~3.8% intl","https://flutterwave.com/pricing","https://dashboard.flutterwave.com/signup"),
 ("Azure OpenAI","Paid","Growth (feature)","per 1M tokens","GPT-4o ~$2.50/$10; mini ~$0.15/$0.60","https://azure.microsoft.com/en-us/pricing/details/cognitive-services/openai-service/","https://azure.microsoft.com/free"),
 ("Azure AI Vision (OCR)","Free tier / Paid","Growth (feature)","per 1k pages","free 5k/mo; ~$1-1.50/1k","https://azure.microsoft.com/en-us/pricing/details/cognitive-services/computer-vision/","https://azure.microsoft.com/free"),
 ("Autodesk APS / Forge","Paid","Enterprise (opt)","flex tokens","pay-as-you-go credits","https://aps.autodesk.com/pricing","https://aps.autodesk.com"),
 ("IfcOpenShell (3D conv.)","Free","Pilot","self-host","$0","https://ifcopenshell.org/","https://ifcopenshell.org/"),
 ("Sentry","Free tier / Paid","Scale","events/users","free / $26/mo Team","https://sentry.io/pricing/","https://sentry.io/signup/"),
 ("Mapbox","Free tier / Paid","Pilot (site)","map loads","free 50k/mo","https://www.mapbox.com/pricing","https://account.mapbox.com/auth/signup/"),
 ("Okta SSO","Paid","Enterprise","per user/mo","~$2/user/mo","https://www.okta.com/pricing/","https://www.okta.com/free-trial/"),
 ("Microsoft Entra ID P1","Paid","Enterprise","per user/mo","~$6/user/mo","https://www.microsoft.com/security/business/microsoft-entra-pricing","https://azure.microsoft.com/free"),
 ("Seq (logs)","Free single / Paid","Scale","per seat","free 1-user; Team paid","https://datalust.co/pricing","https://datalust.co"),
 ("Slack (webhooks)","Free tier / Paid","Growth (opt)","per user/mo","free / ~$7.25/user","https://slack.com/pricing","https://slack.com/get-started"),
 ("Apple Developer","Paid","when native iOS","flat/yr","$99/yr","https://developer.apple.com/programs/","https://developer.apple.com/programs/enroll/"),
 ("Google Play Console","Paid","when native Android","one-time","$25 once","https://play.google.com/console/signup","https://play.google.com/console/signup"),
 ("Expo EAS","Free tier / Paid","when native apps","build mins/seats","free / $99/mo Production","https://expo.dev/pricing","https://expo.dev/signup"),
 ("Domain registrar","Paid","Pilot","per domain/yr","~$12/yr","https://www.cloudflare.com/products/registrar/","https://dash.cloudflare.com/sign-up"),
]
r=4
for row in cat:
    for i,v in enumerate(row,1):
        cell=sc.cell(row=r,column=i,value=v); cell.border=border; cell.alignment=wrap
        if i==1: cell.font=bold
        if i==2 and v=="Free": cell.font=Font(bold=True,color=GREEN)
        if i==2 and v=="Paid": cell.font=Font(bold=True,color=RED)
        if i in (6,7) and isinstance(v,str) and v.startswith("http"): cell.hyperlink=v; cell.font=link_font
    r+=1
for col,w in zip("ABCDEFG",[26,22,16,18,30,58,50]): sc.column_dimensions[col].width=w
sc.freeze_panes="A4"

# ============ SHEET 5: Deploy by Tier ============
dp=wb.create_sheet("Deploy by Tier"); dp.sheet_properties.tabColor="546E7A"
dp["A1"]="ARCHITECTURE & PROVISIONING BY TIER"; dp["A1"].font=Font(bold=True,size=14,color=NAVY)
dp.merge_cells("A1:C1")
for i,h in enumerate(["Tier","Architecture","Provisioning checklist"],1): dp.cell(row=3,column=i,value=h)
hrow(dp,3,3)
dep=[
 ("Pilot","Single VPS. Full docker-compose (api+worker+postgres+redis). Local-disk storage. Caddy for TLS.",
  "1) Create VPS + DNS A-record  2) clone repo  3) .env: set DB_PASSWORD + JWT_KEY only  4) docker compose up -d (no extra profiles)  5) Caddy/nginx TLS  6) set Cors__Origins + EXPO_PUBLIC_API_BASE  7) send link."),
 ("Starter","Single larger VPS. Storage moved to R2/Spaces. Postmark email. Stripe/Flutterwave keys added.",
  "Resize VPS to 4 vCPU/16 GB; set Storage__S3__* to R2; set Smtp__*; add Billing__Stripe__* + webhook; keep Postgres/Redis in-box."),
 ("Growth","Managed Postgres + managed Redis. App + worker as separate containers/nodes. Nightly backups.",
  "Provision managed PG (point ConnectionStrings__Default at it) + managed Redis (Redis__Connection); run worker with PLANSCAPE_ROLE=worker; enable backups; (opt) set Azure OpenAI/Vision keys + ModelConverter."),
 ("Scale","HA: PG primary+standby, 2x app nodes behind LB, PgBouncer (--profile pool), Redis cluster. Observability on.",
  "Add standby + failover; load balancer; docker compose --profile pool (PgBouncer); --profile observability (Prometheus/Grafana); Sentry DSN; scale storage to 5 TB; rate-limit via Redis."),
 ("Enterprise","Multi-node, multi-AZ. Dedicated DB. SSO/SAML/SCIM. APS conversion. Managed SIEM/logs. AV scanning.",
  "Configure per-tenant SsoConfig (OIDC/SAML/SCIM); ModelConverter__Provider=aps + APS creds; --profile av (ClamAV); ship logs to managed Seq/Elastic; 24/7 monitoring + on-call; security review."),
]
r=4
for tier,arch,chk in dep:
    dp.cell(row=r,column=1,value=tier).font=bold
    dp.cell(row=r,column=2,value=arch); dp.cell(row=r,column=3,value=chk)
    for c in range(1,4): dp.cell(row=r,column=c).border=border; dp.cell(row=r,column=c).alignment=wrap
    r+=1
for col,w in zip("ABC",[14,52,72]): dp.column_dimensions[col].width=w

out="/home/user/STINGTOOLS/Planscape_Tier_Budget.xlsx"
wb.save(out)
print("Saved",out); print("Sheets:",wb.sheetnames)
