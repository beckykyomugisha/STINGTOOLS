#!/usr/bin/env python3
"""generate_paragraph_xlsx.py — emit a per-family tag config XLSX whose
T4..T10 row blocks are populated from a named preset in
StingTools/Data/PARAGRAPH_PRESETS.json.

Used to produce switchable preset-mode workbooks alongside the universal
default (e.g. the Handover XLSX vs the Design & Construction XLSX) without
regenerating the underlying CSVs. Family list, header, and styling are
preserved from a source XLSX template.

Usage:
    python3 generate_paragraph_xlsx.py <preset_key> [<source_xlsx>] [<out_xlsx>]

Example:
    python3 generate_paragraph_xlsx.py DesignConstruction
    python3 generate_paragraph_xlsx.py Handover
"""
import json
import os
import re
import shutil
import sys
from datetime import datetime

from openpyxl import load_workbook

REPO_ROOT = os.path.dirname(os.path.abspath(__file__))
PRESETS_JSON = os.path.join(REPO_ROOT, 'StingTools', 'Data', 'PARAGRAPH_PRESETS.json')
DEFAULT_SOURCE = os.path.join(
    REPO_ROOT, '20260423_sting_tag_config_142_v6_0_per_family.xlsx')

TIERS = ('T4', 'T5', 'T6', 'T7', 'T8', 'T9', 'T10')


def load_preset(key):
    with open(PRESETS_JSON, encoding='utf-8') as f:
        data = json.load(f)
    if key not in data['presets']:
        raise SystemExit(f'Preset "{key}" not in PARAGRAPH_PRESETS.json. '
                         f'Available: {list(data["presets"])}')
    return data['presets'][key]


def patch_workbook(src, dst, preset_key, preset):
    shutil.copy(src, dst)
    wb = load_workbook(dst)

    tiers = preset['tiers']
    seq_no = None  # column 0 numeric counter, reset per family
    current_tier = None
    rows_consumed = 0
    families_touched = 0

    sheet_names = wb.sheetnames
    for sn in sheet_names:
        ws = wb[sn]
        # Reset per-sheet state
        current_tier = None
        rows_consumed = 0
        for row in ws.iter_rows(min_row=1):
            c0 = row[0].value if len(row) > 0 else None
            c1 = row[1].value if len(row) > 1 else None

            # Tier banner: "━━━ Tier T4 ━━━" (v6) or "T4 — Label" (v5)
            if isinstance(c0, str):
                m = re.search(r'Tier\s+(T(?:[4-9]|10))', c0) \
                    or re.match(r'^(T(?:[4-9]|10))\s+—', c0)
                if m:
                    current_tier = m.group(1)
                    rows_consumed = 0
                    new_label = tiers.get(current_tier, {}).get('label', '')
                    if new_label:
                        if c0.startswith('━'):
                            row[0].value = f'━━━ Tier {current_tier} — {new_label} ━━━'
                        else:
                            row[0].value = f'{current_tier} — {new_label}'
                    continue

            # Tier data row: c1 in TIERS
            if isinstance(c1, str) and c1 in TIERS:
                t = c1
                src_rows = tiers.get(t, {}).get('rows', [])
                if rows_consumed < len(src_rows):
                    r = src_rows[rows_consumed]
                    label = tiers.get(t, {}).get('label', '')
                    tier_num = t[1:]
                    set_if_in_range(row, 2, r.get('parameter', ''))
                    set_if_in_range(row, 3, r.get('prefix', ''))
                    set_if_in_range(row, 4, r.get('suffix', ''))
                    set_if_in_range(row, 6, '✓' if r.get('break') else '')
                    set_if_in_range(row, 9,
                        f'Show {t} - {label} - {r.get("parameter","")}')
                    set_if_in_range(row, 10,
                        f'if(TAG_PARA_STATE_{tier_num}_BOOL, {r.get("parameter","")}, "")')
                    set_if_in_range(row, 11, r.get('style', 'NOM'))
                    set_if_in_range(row, 12, r.get('color', 'GREY'))
                    set_if_in_range(row, 13, r.get('size', 2.0))
                    rows_consumed += 1
                else:
                    # Preset has fewer rows than the template — blank the surplus
                    for ci in (2, 3, 4, 9, 10):
                        set_if_in_range(row, ci, '')

            # Schema banner (row 1)
            if isinstance(c0, str) and c0.startswith('#SCHEMA_VERSION='):
                row[0].value = (
                    f'#SCHEMA_VERSION=6.0,PRESET={preset_key},'
                    f'CREATED={datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")}')
        families_touched += 1

    wb.save(dst)
    return families_touched


def set_if_in_range(row, col_idx, value):
    if col_idx < len(row):
        row[col_idx].value = value


def main(argv):
    if len(argv) < 2:
        print(__doc__)
        sys.exit(1)
    preset_key = argv[1]
    source = argv[2] if len(argv) > 2 else DEFAULT_SOURCE
    if not os.path.exists(source):
        raise SystemExit(f'Source XLSX not found: {source}')
    if len(argv) > 3:
        dst = argv[3]
    else:
        stamp = datetime.utcnow().strftime('%Y%m%d')
        dst = os.path.join(
            REPO_ROOT,
            f'{stamp}_sting_tag_config_v6_0_{preset_key}.xlsx')

    preset = load_preset(preset_key)
    families = patch_workbook(source, dst, preset_key, preset)
    print(f'OK — wrote {dst}')
    print(f'   preset={preset_key}  families/sheets={families}')


if __name__ == '__main__':
    main(sys.argv)
