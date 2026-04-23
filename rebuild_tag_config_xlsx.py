#!/usr/bin/env python3
"""rebuild_tag_config_xlsx.py — consume the 4 regenerated
STING_TAG_CONFIG_v5_0_*.csv files and emit a colour-banded workbook:

  20260423_sting_tag_config_142_v6_0_per_family.xlsx

One worksheet per Tag Family (142). Each sheet shows T1..T10 rows
verbatim from the CSVs (OMITted tiers are already absent in the source)
with one banner row per tier and per-tier colour bands mirroring the
existing workbook style.

Schema header bumped to SCHEMA_VERSION=6.0.
"""
import os
import re
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPO_ROOT = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(REPO_ROOT, 'StingTools', 'Data')

CSVS = [
    os.path.join(DATA, 'STING_TAG_CONFIG_v5_0_ARCH.csv'),
    os.path.join(DATA, 'STING_TAG_CONFIG_v5_0_GEN.csv'),
    os.path.join(DATA, 'STING_TAG_CONFIG_v5_0_MEP.csv'),
    os.path.join(DATA, 'STING_TAG_CONFIG_v5_0_STR.csv'),
]

# Tier colour bands (mirrors the CSV Style/Color defaults per tier)
TIER_BANDS = {
    'T1':  '1F4E78',  # deep blue
    'T2':  'D9E1F2',  # light blue
    'T3':  'E2EFDA',  # light green
    'T4':  'DDEBF7',  # commissioning — light blue
    'T5':  'F4CCCC',  # cost — light purple/pink
    'T6':  'D9EAD3',  # carbon — light green
    'T7':  'FCE5CD',  # fabrication — light orange
    'T8':  'F4CCCC',  # clash — light red
    'T9':  'D9D9D9',  # as-built — light grey
    'T10': 'EFEFEF',  # compliance — very light grey
}

WARN_FILL   = 'F8CBAD'  # warnings band — orange-red
HEADER_FILL = '305496'  # dark blue for column header
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
BANNER_FONT = Font(name='Calibri', bold=True, color='1F1F1F', size=11)
CELL_FONT   = Font(name='Calibri', size=10)
THIN        = Side(border_style='thin', color='BFBFBF')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_csv_line(line):
    """Handle quoted fields with embedded commas."""
    cols = []
    cur = []
    in_q = False
    i = 0
    while i < len(line):
        c = line[i]
        if in_q:
            if c == '"' and i+1 < len(line) and line[i+1] == '"':
                cur.append('"')
                i += 2
                continue
            if c == '"':
                in_q = False
                i += 1
                continue
            cur.append(c)
            i += 1
        else:
            if c == '"':
                in_q = True
                i += 1
                continue
            if c == ',':
                cols.append(''.join(cur))
                cur = []
                i += 1
                continue
            cur.append(c)
            i += 1
    cols.append(''.join(cur))
    return cols


def extract_families():
    """Return list of (discipline, family_name, tier_rows, warning_rows, category)."""
    result = []
    for path in CSVS:
        disc = re.search(r'_(ARCH|GEN|MEP|STR)\.csv$', path).group(1)
        with open(path, encoding='utf-8') as f:
            text = f.read()
        blocks = re.split(r'(?=^Tag Family #\d+: )', text, flags=re.M)
        for b in blocks:
            if not b.startswith('Tag Family #'):
                continue
            lines = b.splitlines()
            header = lines[0]
            name_m = re.match(r'Tag Family #\d+: (.+)$', header)
            if not name_m:
                continue
            name = name_m.group(1).strip()
            cat = '?'
            if len(lines) > 1:
                cm = re.search(r'Category:\s*([^|•]+?)(?:\s*[\|•]|\s*$)', lines[1])
                if cm:
                    cat = cm.group(1).strip()
            tier_rows = []
            warn_rows = []
            in_warn = False
            for ln in lines[2:]:
                if ln.startswith('⚠ WARNING PARAMETERS'):
                    in_warn = True
                    continue
                if ln.startswith('#,') or ln.startswith('Tag Family #') or not ln.strip():
                    continue
                cols = parse_csv_line(ln)
                if in_warn:
                    if cols and cols[0].isdigit():
                        warn_rows.append(cols)
                    continue
                if cols and cols[0].isdigit():
                    tier_rows.append(cols)
            result.append((disc, name, cat, tier_rows, warn_rows))
    return result


def sheet_name_safe(name):
    """Excel sheet names: ≤31 chars, no / \\ ? * [ ] :"""
    s = re.sub(r'[\\/\?\*\[\]\:]', '_', name)
    # Strip "STING - " prefix to save chars
    s = s.replace('STING - ', '')
    return s[:31]


def populate_sheet(ws, disc, name, category, tier_rows, warn_rows):
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    c = ws.cell(row=1, column=1, value=f"{name}   [{disc}] • Category: {category}")
    c.font = Font(name='Calibri', bold=True, size=14, color='1F4E78')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    # Column header
    headers = ['#', 'Tier', 'Parameter', 'Prefix', 'Suffix', 'Spc', 'Brk',
               'Discipline', 'Type', 'Name', 'Formula', 'Style', 'Color', 'Size', 'Box', 'Arrow']
    for col, h in enumerate(headers, start=1):
        cc = ws.cell(row=3, column=col, value=h)
        cc.font = HEADER_FONT
        cc.fill = PatternFill('solid', fgColor=HEADER_FILL)
        cc.alignment = Alignment(horizontal='center', vertical='center')
        cc.border = BORDER

    r = 4
    current_tier = None
    for row in tier_rows:
        tier = row[1] if len(row) > 1 else ''
        if tier != current_tier:
            # Banner row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=16)
            band_color = TIER_BANDS.get(tier, 'FFFFFF')
            banner = ws.cell(row=r, column=1, value=f"━━━ Tier {tier} ━━━")
            banner.font = BANNER_FONT
            banner.fill = PatternFill('solid', fgColor=band_color)
            banner.alignment = Alignment(horizontal='left', vertical='center')
            banner.border = BORDER
            current_tier = tier
            r += 1
        for col, val in enumerate(row[:16], start=1):
            cc = ws.cell(row=r, column=col, value=val)
            cc.font = CELL_FONT
            cc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            cc.fill = PatternFill('solid', fgColor=TIER_BANDS.get(tier, 'FFFFFF'))
            cc.border = BORDER
        r += 1

    # Warning block
    if warn_rows:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        wb = ws.cell(row=r, column=1, value="⚠ WARNING PARAMETERS")
        wb.font = Font(name='Calibri', bold=True, color='9C0006', size=11)
        wb.fill = PatternFill('solid', fgColor=WARN_FILL)
        wb.alignment = Alignment(horizontal='left', vertical='center')
        wb.border = BORDER
        r += 1
        # Warning column header
        wheaders = ['#', 'Sev', 'Warning Parameter', 'Threshold', 'Standard', 'Dis.', 'Type', 'Name', 'Formula']
        for col, h in enumerate(wheaders, start=1):
            cc = ws.cell(row=r, column=col, value=h)
            cc.font = HEADER_FONT
            cc.fill = PatternFill('solid', fgColor=HEADER_FILL)
            cc.alignment = Alignment(horizontal='center', vertical='center')
            cc.border = BORDER
        r += 1
        for wrow in warn_rows:
            for col, val in enumerate(wrow[:9], start=1):
                cc = ws.cell(row=r, column=col, value=val)
                cc.font = CELL_FONT
                cc.alignment = Alignment(horizontal='left', vertical='center')
                cc.fill = PatternFill('solid', fgColor=WARN_FILL)
                cc.border = BORDER
            r += 1

    # Column widths
    widths = [5, 6, 32, 10, 10, 5, 5, 10, 8, 38, 48, 10, 10, 6, 6, 6]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'+chr(64 + i - 26)].width = w


def main():
    families = extract_families()
    print(f'Loaded {len(families)} families from 4 CSVs')

    wb = Workbook()
    wb.remove(wb.active)

    # --- Index sheet ---
    idx = wb.create_sheet('_Index', 0)
    idx.cell(row=1, column=1,
             value='STING Tag Config v6.0 — Per-family T4..T10 tier variation')
    idx.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=14)
    idx.cell(row=2, column=1,
             value='SCHEMA_VERSION=6.0,NOTES=per_family_tier_variation_with_standards_cited_and_reuse_check')
    idx.cell(row=2, column=1).font = Font(name='Calibri', italic=True, size=10, color='6A6A6A')
    headers = ['#', 'Discipline', 'Family', 'Category', 'Tier rows', 'Warnings']
    for col, h in enumerate(headers, start=1):
        cc = idx.cell(row=4, column=col, value=h)
        cc.font = HEADER_FONT
        cc.fill = PatternFill('solid', fgColor=HEADER_FILL)
        cc.border = BORDER
    for i, (disc, name, cat, tier_rows, warn_rows) in enumerate(families, start=1):
        idx.cell(row=4+i, column=1, value=i)
        idx.cell(row=4+i, column=2, value=disc)
        idx.cell(row=4+i, column=3, value=name)
        idx.cell(row=4+i, column=4, value=cat)
        idx.cell(row=4+i, column=5, value=len(tier_rows))
        idx.cell(row=4+i, column=6, value=len(warn_rows))
    for i, w in enumerate([5, 10, 44, 40, 12, 12], start=1):
        idx.column_dimensions[chr(64 + i)].width = w

    # --- Per-family sheets ---
    seen_names = set()
    for disc, name, cat, tier_rows, warn_rows in families:
        sn = sheet_name_safe(name)
        orig = sn
        # Dedupe sheet names (case-insensitive)
        k = 2
        while sn.lower() in seen_names:
            sn = (orig[:28] + f"_{k}")[:31]
            k += 1
        seen_names.add(sn.lower())
        ws = wb.create_sheet(sn)
        populate_sheet(ws, disc, name, cat, tier_rows, warn_rows)

    out = os.path.join(REPO_ROOT, '20260423_sting_tag_config_142_v6_0_per_family.xlsx')
    wb.save(out)
    print(f'Wrote {out}: {len(families)+1} sheets (1 index + {len(families)} families)')

if __name__ == '__main__':
    sys.exit(main())
