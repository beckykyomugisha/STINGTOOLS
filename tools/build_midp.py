# -*- coding: utf-8 -*-
"""Build the KUT Master Information Delivery Plan as a corporate .xlsx workbook.

Source content: GUIDES/KUT_MIDP_TEMPLATE.csv, with the tooling references removed
(this is a project document), the originator column reset pending the code
register, and the Stage 3.1 asset-data capture rows added.

Palette and register conventions match the issued Word documents produced by
tools/corporate_docx.py.
"""
import datetime
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
import kut_docs_lib as K                                          # noqa: E402
from midp_schema import COLS, LIST_COL, LISTS, MIDP_VALIDATED, TIDP_VALIDATED   # noqa: E402

OUT = 'KUT_Master_Information_Delivery_Plan.xlsx'

NAVY = '1F3654'
SLATE = '444F5C'
BAND = 'E8EDF3'
SHADE = 'F4F6F9'
LINE = 'C6CEDA'
RED = 'F8D7DA'
AMBER = 'FFF0CC'
GREEN = 'DDF0DD'

thin = Side(style='thin', color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

title_f = Font(name='Calibri', size=22, bold=True, color=NAVY)
eyebrow_f = Font(name='Calibri', size=10, bold=True, color=SLATE)
h_f = Font(name='Calibri', size=11, bold=True, color=NAVY)
lbl_f = Font(name='Calibri', size=9, bold=True, color=SLATE)
body_f = Font(name='Calibri', size=9.5)
small_f = Font(name='Calibri', size=8.5, italic=True, color=SLATE)
hdr_fill = PatternFill('solid', fgColor=BAND)
shade_fill = PatternFill('solid', fgColor=SHADE)

# Validation lists and register columns live in tools/midp_schema.py so that
# tools/merge_tidp.py validates a returned TIDP against exactly what this
# workbook offered. See that module for why.

# ── register rows ───────────────────────────────────────────────────────────
# Ref, Discipline, Deliverable, Type, Stage, LOD, Format, Suit, State, Month, Responsible, TIDP, Notes
R = [
    ('Z-000', 'Information Management', 'BIM Execution Plan', 'Document', 'Mobilisation', 'n/a', 'DOCX/PDF', 'A1', 'Published', 'M0', 'Information Manager', 'TIDP-Z', 'Baselined at mobilisation'),
    ('Z-001', 'Information Management', 'Master Information Delivery Plan', 'Document', 'Mobilisation', 'n/a', 'XLSX', 'A1', 'Published', 'M0', 'Information Manager', 'TIDP-Z', 'Aggregated from all TIDPs'),
    ('Z-002', 'Information Management', 'Project Delivery Playbook', 'Document', 'Mobilisation', 'n/a', 'DOCX/PDF', 'A1', 'Published', 'M0', 'Information Manager', 'TIDP-Z', ''),
    ('Z-003', 'Information Management', 'Responsibility matrix (RACI)', 'Document', 'Mobilisation', 'n/a', 'XLSX', 'A1', 'Published', 'M0', 'Information Manager', 'TIDP-Z', ''),
    ('Z-004', 'Information Management', 'Project template, family library and title blocks', 'Document', 'Mobilisation', 'n/a', 'RVT/RFA', 'A1', 'Published', 'M0', 'Information Manager', 'TIDP-Z', 'Issued to every task team'),
    ('Z-005', 'Information Management', 'Originator code register', 'Document', 'Mobilisation', 'n/a', 'XLSX', 'A1', 'Published', 'M0', 'Appointing Party', 'TIDP-Z', 'Required before any container is numbered'),
    ('Z-007', 'Information Management', 'Standards reconciliation schedule (BEP section 4.1.3)', 'Document', 'Mobilisation', 'n/a', 'DOCX/PDF', 'S3', 'Shared', 'M0', 'All disciplines', 'TIDP-Z', 'Opened at mobilisation; CLOSED before Technical Design. An open row means two parties may design to different bases'),
    ('Z-008', 'Information Management', 'Shared coordinates seed model', 'Model', 'Mobilisation', 'n/a', 'RVT', 'A1', 'Published', 'M0', 'Information Manager', 'TIDP-Z', 'Survey point, project base point, true north, levels and grids. Every model acquires coordinates from it'),
    ('Z-009', 'Information Management', 'Level and grid register', 'Schedule', 'Mobilisation', 'n/a', 'XLSX/PDF', 'A1', 'Published', 'M0', 'Lead Appointed Party', 'TIDP-Z', 'Level codes appear in every container name'),
    ('Z-011', 'Information Management', 'Software and version register', 'Schedule', 'Mobilisation', 'n/a', 'XLSX', 'A1', 'Published', 'M0', 'All parties', 'TIDP-Z', 'One authoring version project-wide for the duration of a stage'),
    ('Z-012', 'Information Management', 'CDE folder structure and permission matrix', 'Document', 'Mobilisation', 'n/a', 'PDF', 'A1', 'Published', 'M0', 'Information Manager', 'TIDP-Z', ''),
    ('Z-013', 'Information Management', 'Information classification and access schedule', 'Document', 'Mobilisation', 'n/a', 'PDF', 'A1', 'Published', 'M0', 'Appointing Party', 'TIDP-Z', 'Restricted spaces and systems must be classified BEFORE modelling begins'),
    ('Z-014', 'Information Management', 'Clash matrix and tolerance schedule', 'Document', 'Mobilisation', 'n/a', 'PDF', 'S3', 'Shared', 'M1', 'MEP and structural leads', 'TIDP-Z', 'Required before the first coordination cycle of Stage 2.2'),
    ('Z-015', 'Information Management', 'IFC export configuration', 'Document', 'Mobilisation', 'n/a', 'PDF/JSON', 'A1', 'Published', 'M0', 'Information Manager', 'TIDP-Z', 'Issued centrally, or every consultant IFC differs'),
    ('Z-016', 'Information Management', 'Shared parameter file', 'Document', 'Mobilisation', 'n/a', 'TXT', 'A1', 'Published', 'M0', 'Information Manager', 'TIDP-Z', 'Central ownership; a locally added parameter has a different GUID'),
    ('Z-017', 'Information Management', 'Capability and capacity assessment', 'Report', 'Mobilisation', 'n/a', 'XLSX', 'S3', 'Shared', 'M1', 'Lead Appointed Party', 'TIDP-Z', 'ISO 19650-2; the Stage 0 test model is its practical form'),
    ('Z-006', 'All disciplines', 'Task Information Delivery Plan', 'Document', 'Mobilisation', 'n/a', 'XLSX', 'S3', 'Shared', 'M0', 'Task Team Managers', 'TIDP-ALL', 'One per appointed party'),
    ('Z-010', 'Information Management', 'Federated coordination model', 'Model', '2.1 Deliverable A', '200', 'NWC/IFC', 'S1', 'Shared', 'M1', 'BIM Coordinator', 'TIDP-Z', 'First federation'),
    ('A-100', 'Architecture', 'Architectural model', 'Model', '2.1 Deliverable A', '200', 'RVT/IFC', 'S2', 'Shared', 'M1', 'Architecture lead', 'TIDP-A', 'Massing and generic systems; rooms placed'),
    ('S-100', 'Structure', 'Structural model', 'Model', '2.1 Deliverable A', '200', 'RVT/IFC', 'S2', 'Shared', 'M1', 'Structural lead', 'TIDP-S', 'Indicative frame and foundations'),
    ('M-100', 'Mechanical', 'Mechanical model', 'Model', '2.1 Deliverable A', '200', 'RVT/IFC', 'S2', 'Shared', 'M1', 'MEP lead', 'TIDP-M', 'Plant space allocation, primary routes'),
    ('E-100', 'Electrical', 'Electrical model', 'Model', '2.1 Deliverable A', '200', 'RVT/IFC', 'S2', 'Shared', 'M1', 'MEP lead', 'TIDP-E', 'Plant space allocation, primary routes'),
    ('P-100', 'Public Health', 'Plumbing and drainage model', 'Model', '2.1 Deliverable A', '200', 'RVT/IFC', 'S2', 'Shared', 'M1', 'MEP lead', 'TIDP-P', 'Primary routes'),
    ('G-100', 'Civil and Site', 'Site model and levels', 'Model', '2.1 Deliverable A', '200', 'RVT/IFC', 'S2', 'Shared', 'M1', 'Civil lead', 'TIDP-G', 'Site levels, access, drainage strategy'),
    ('Z-110', 'Information Management', 'Basis of design report', 'Document', '2.1 Deliverable A', '200', 'DOCX/PDF', 'S3', 'Shared', 'M1', 'Lead Appointed Party', 'TIDP-Z', 'Design intent'),
    ('Z-111', 'Information Management', 'Area schedule reconciled to the brief', 'Schedule', '2.1 Deliverable A', '200', 'XLSX/PDF', 'S3', 'Shared', 'M1', 'Information Manager', 'TIDP-Z', 'Gate condition'),
    ('Z-112', 'Information Management', 'Deliverable A gate pack', 'Document', '2.1 Deliverable A', '200', 'PDF', 'S4', 'Shared', 'M1', 'Information Manager', 'TIDP-Z', 'Contents per the playbook, Appendix B'),
    ('A-200', 'Architecture', 'Architectural model', 'Model', '2.2 Deliverable B', '300', 'RVT/IFC', 'S2', 'Shared', 'M4', 'Architecture lead', 'TIDP-A', 'Real geometry, correctly located'),
    ('A-201', 'Architecture', 'General arrangement plans, sections and elevations (50%)', 'Drawing', '2.2 Deliverable B', '300', 'PDF/DWG', 'S2', 'Shared', 'M4', 'Architecture lead', 'TIDP-A', ''),
    ('A-202', 'Architecture', 'Door and window schedules', 'Schedule', '2.2 Deliverable B', '300', 'PDF/XLSX', 'S2', 'Shared', 'M4', 'Architecture lead', 'TIDP-A', ''),
    ('A-203', 'Architecture', 'Room data sheets (key spaces)', 'Room data sheet', '2.2 Deliverable B', '300', 'PDF', 'S2', 'Shared', 'M4', 'Architecture lead', 'TIDP-A', ''),
    ('A-204', 'Architecture', 'Existing conditions and removals plan', 'Drawing', '2.2 Deliverable B', '300', 'PDF/DWG', 'S2', 'Shared', 'M4', 'Architecture lead', 'TIDP-A', 'Demolition classification is a manual task; see BEP 10.4'),
    ('S-200', 'Structure', 'Structural model', 'Model', '2.2 Deliverable B', '300', 'RVT/IFC', 'S2', 'Shared', 'M4', 'Structural lead', 'TIDP-S', 'Sized members; penetrations coordinated'),
    ('S-201', 'Structure', 'General arrangement drawings (50%)', 'Drawing', '2.2 Deliverable B', '300', 'PDF/DWG', 'S2', 'Shared', 'M4', 'Structural lead', 'TIDP-S', ''),
    ('M-200', 'Mechanical', 'Mechanical model', 'Model', '2.2 Deliverable B', '300', 'RVT/IFC', 'S2', 'Shared', 'M4', 'MEP lead', 'TIDP-M', 'Real equipment; risers fixed'),
    ('M-201', 'Mechanical', 'HVAC layouts (50%)', 'Drawing', '2.2 Deliverable B', '300', 'PDF/DWG', 'S2', 'Shared', 'M4', 'MEP lead', 'TIDP-M', ''),
    ('E-200', 'Electrical', 'Electrical model', 'Model', '2.2 Deliverable B', '300', 'RVT/IFC', 'S2', 'Shared', 'M4', 'MEP lead', 'TIDP-E', ''),
    ('E-201', 'Electrical', 'Power and lighting layouts (50%)', 'Drawing', '2.2 Deliverable B', '300', 'PDF/DWG', 'S2', 'Shared', 'M4', 'MEP lead', 'TIDP-E', ''),
    ('P-200', 'Public Health', 'Plumbing and drainage model', 'Model', '2.2 Deliverable B', '300', 'RVT/IFC', 'S2', 'Shared', 'M4', 'MEP lead', 'TIDP-P', ''),
    ('FP-200', 'Fire Protection', 'Fire strategy and suppression layout', 'Model/Document', '2.2 Deliverable B', '300', 'RVT/PDF', 'S2', 'Shared', 'M4', 'Fire lead', 'TIDP-FP', ''),
    ('Z-210', 'Information Management', 'Coordination report', 'Report', '2.2 Deliverable B', '300', 'PDF/BCF', 'S2', 'Shared', 'M4', 'BIM Coordinator', 'TIDP-Z', 'No unresolved high priority clashes'),
    ('Z-211', 'Information Management', 'Federated model', 'Model', '2.2 Deliverable B', '300', 'NWC/IFC', 'S2', 'Shared', 'M4', 'BIM Coordinator', 'TIDP-Z', ''),
    ('Z-212', 'QS / Cost', 'Bill of quantities (preliminary)', 'Schedule', '2.2 Deliverable B', '300', 'XLSX', 'S2', 'Shared', 'M4', 'Quantity Surveyor', 'TIDP-Q', ''),
    ('FF-200', 'FF&E', 'FF&E and finishes register (first exchange)', 'Schedule', '2.2 Deliverable B', '300', 'CSV/XLSX', 'S2', 'Shared', 'M4', 'Interior Designer', 'TIDP-FF', 'Matched on room number'),
    ('Z-213', 'Information Management', 'Deliverable B gate pack', 'Document', '2.2 Deliverable B', '300', 'PDF', 'S4', 'Shared', 'M4', 'Information Manager', 'TIDP-Z', ''),
    ('A-300', 'Architecture', 'Architectural model', 'Model', '2.3 Deliverable C', '350', 'RVT/IFC', 'S4', 'Shared', 'M8', 'Architecture lead', 'TIDP-A', 'Interfaces resolved'),
    ('A-301', 'Architecture', 'Full drawing set (100%) and details', 'Drawing', '2.3 Deliverable C', '350', 'PDF/DWG', 'S4', 'Shared', 'M8', 'Architecture lead', 'TIDP-A', ''),
    ('S-300', 'Structure', 'Structural model', 'Model', '2.3 Deliverable C', '350', 'RVT/IFC', 'S4', 'Shared', 'M8', 'Structural lead', 'TIDP-S', 'Connections resolved'),
    ('S-301', 'Structure', 'Structural drawing set (100%)', 'Drawing', '2.3 Deliverable C', '350', 'PDF/DWG', 'S4', 'Shared', 'M8', 'Structural lead', 'TIDP-S', ''),
    ('M-300', 'Mechanical', 'Mechanical model and drawing set (100%)', 'Model/Drawing', '2.3 Deliverable C', '350', 'RVT/PDF', 'S4', 'Shared', 'M8', 'MEP lead', 'TIDP-M', 'BMS points identified'),
    ('E-300', 'Electrical', 'Electrical model and drawing set (100%)', 'Model/Drawing', '2.3 Deliverable C', '350', 'RVT/PDF', 'S4', 'Shared', 'M8', 'MEP lead', 'TIDP-E', ''),
    ('P-300', 'Public Health', 'Plumbing model and drawing set (100%)', 'Model/Drawing', '2.3 Deliverable C', '350', 'RVT/PDF', 'S4', 'Shared', 'M8', 'MEP lead', 'TIDP-P', ''),
    ('FP-300', 'Fire Protection', 'Fire strategy and model (100%)', 'Model/Document', '2.3 Deliverable C', '350', 'RVT/PDF', 'S4', 'Shared', 'M8', 'Fire lead', 'TIDP-FP', ''),
    ('LV-300', 'Low Voltage', 'Communications and security model', 'Model/Drawing', '2.3 Deliverable C', '350', 'RVT/PDF', 'S4', 'Shared', 'M8', 'LV lead', 'TIDP-LV', ''),
    ('Z-310', 'QS / Cost', 'Bill of quantities (tender)', 'Schedule', '2.3 Deliverable C', '350', 'XLSX', 'S4', 'Shared', 'M8', 'Quantity Surveyor', 'TIDP-Q', ''),
    ('Z-311', 'Information Management', 'Specifications', 'Specification', '2.3 Deliverable C', '350', 'PDF', 'S4', 'Shared', 'M8', 'Lead Appointed Party', 'TIDP-Z', ''),
    ('Z-312', 'Information Management', 'Specification reconciliation report', 'Report', '2.3 Deliverable C', '350', 'XLSX/PDF', 'S4', 'Shared', 'M8', 'Information Manager', 'TIDP-Z', 'Gaps closed or formally accepted'),
    ('FF-300', 'FF&E', 'FF&E specifications and schedule', 'Schedule', '2.3 Deliverable C', '350', 'XLSX/PDF', 'S4', 'Shared', 'M8', 'Interior Designer', 'TIDP-FF', ''),
    ('Z-313', 'Information Management', 'Deliverable C gate pack', 'Document', '2.3 Deliverable C', '350', 'PDF', 'S4', 'Shared', 'M8', 'Information Manager', 'TIDP-Z', ''),
    ('Z-400', 'Information Management', 'Tender documents', 'Document', '2.4 Tender', 'n/a', 'PDF', 'A1', 'Published', 'M9', 'Lead Appointed Party', 'TIDP-Z', ''),
    ('Z-401', 'Information Management', 'Tender query and addendum log', 'Report', '2.4 Tender', 'n/a', 'XLSX', 'S2', 'Shared', 'M10', 'Lead Appointed Party', 'TIDP-Z', ''),
    ('Z-410', 'Information Management', 'Conformed set', 'Drawing', '2.5 Conformed set', '350', 'PDF/DWG', 'A1', 'Published', 'M11', 'Lead Appointed Party', 'TIDP-Z', 'Post-award addenda incorporated'),
    ('ALL-500', 'All disciplines', 'Construction stage models', 'Model', '3.1 Construction', '400', 'RVT/IFC', 'A1', 'Published', 'M12-M43', 'Task teams', 'TIDP-ALL', 'Ongoing; revisions tracked'),
    ('Z-510', 'Information Management', 'Request for information and submittal register', 'Report', '3.1 Construction', '400', 'XLSX', 'S2', 'Shared', 'M12-M43', 'Information Manager', 'TIDP-Z', 'Maintained continuously'),
    ('Z-511', 'Information Management', 'Monthly status report', 'Report', '3.1 Construction', '400', 'PDF', 'S2', 'Shared', 'M12-M43', 'Information Manager', 'TIDP-Z', 'Monthly throughout construction'),
    ('Z-512', 'Contractor', 'As-built capture (progressive)', 'Model', '3.1 Construction', '400', 'RVT/IFC', 'S2', 'Shared', 'M12-M43', 'Contractor', 'TIDP-C', 'Current to within one month'),
    ('Z-513', 'Contractor', 'Asset data capture (tiered schedule, BEP section 14)', 'Schedule', '3.1 Construction', '400', 'XLSX', 'S2', 'Shared', 'M12-M43', 'Contractor', 'TIDP-C', 'Tier A serialised plant, Tier B maintainable devices, Tier C warranted fabric. Required for LOD 500; reported monthly from the first month of construction'),
    ('Z-514', 'Information Management', 'Asset data completeness report', 'Report', '3.1 Construction', '400', 'XLSX/PDF', 'S2', 'Shared', 'M12-M43', 'Information Manager', 'TIDP-Z', 'Monthly, by tier and by volume. A tier below 95 per cent at the Deliverable D gate is a gate failure'),
    ('M-520', 'Mechanical', 'Commissioning point list', 'Schedule', '3.1 Construction', '400', 'CSV', 'S2', 'Shared', 'M40', 'MEP lead', 'TIDP-M', 'Produced from the model for the controls contractor'),
    ('FF-600', 'FF&E', 'FF&E final schedule and procurement record', 'Schedule', '3.2 FF&E', '400', 'XLSX', 'A1', 'Published', 'M43', 'Interior Designer', 'TIDP-FF', 'Reconciled item by item'),
    ('A-700', 'Architecture', 'As-built architectural model', 'Model', '3.3 Deliverable D', '500', 'RVT/IFC', 'A1', 'Published', 'M45', 'Architecture lead', 'TIDP-A', 'Verified as-built'),
    ('S-700', 'Structure', 'As-built structural model', 'Model', '3.3 Deliverable D', '500', 'RVT/IFC', 'A1', 'Published', 'M45', 'Structural lead', 'TIDP-S', ''),
    ('M-700', 'Mechanical', 'As-built MEP model', 'Model', '3.3 Deliverable D', '500', 'RVT/IFC', 'A1', 'Published', 'M45', 'MEP lead', 'TIDP-M', 'Includes Tier A and Tier B asset data'),
    ('Z-700', 'Information Management', 'COBie handover dataset', 'Schedule', '3.3 Deliverable D', '500', 'XLSX', 'A1', 'Published', 'M45', 'Information Manager', 'TIDP-Z', 'COBie 2.4'),
    ('Z-701', 'Information Management', 'Operation and maintenance / asset data pack', 'Document', '3.3 Deliverable D', '500', 'PDF', 'A1', 'Published', 'M45', 'Information Manager', 'TIDP-Z', ''),
    ('Z-702', 'Information Management', 'Reconciled building management system point register', 'Schedule', '3.3 Deliverable D', '500', 'CSV/PDF', 'A1', 'Published', 'M45', 'Information Manager', 'TIDP-Z', 'Model reconciled to the live station'),
    ('Z-703', 'Information Management', 'Deliverable D gate pack', 'Document', '3.3 Deliverable D', '500', 'PDF', 'A1', 'Published', 'M45', 'Information Manager', 'TIDP-Z', ''),
    ('Z-704', 'Information Management', 'Final transmittal and project archive', 'Document', '3.3 Deliverable D', '500', 'PDF', 'A1', 'Archived', 'M45', 'Information Manager', 'TIDP-Z', ''),
]


def style_header(ws, row, ncols):
    for i in range(1, ncols + 1):
        c = ws.cell(row=row, column=i)
        c.font = Font(name='Calibri', size=9, bold=True, color=NAVY)
        c.fill = hdr_fill
        c.border = box
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30


# ═══ 1. Cover ═══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Cover'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 76

ws['B3'] = 'KAMPALA UGANDA TEMPLE'
ws['B3'].font = eyebrow_f
ws['B5'] = 'Master Information Delivery Plan'
ws['B5'].font = title_f
ws['B7'] = ('The aggregated schedule of information deliverables for the project, compiled from the Task '
            'Information Delivery Plan of every appointed party.')
ws['B7'].font = Font(name='Calibri', size=11, color=SLATE)
ws['B7'].alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells('B7:C7')
ws.row_dimensions[7].height = 32

meta = [
    ('Document reference', 'KUT-SMB-ZZ-ZZ-SC-Z-0001'),
    ('Revision', 'P01'),
    ('Status / suitability', 'A1 — Authorised for use'),
    ('Prepared by', 'Symbion Consulting Group Studios (Information Manager)'),
    ('Baselined', '[FILL — date]'),
    ('Last updated', '[FILL — date]'),
    ('Update frequency', 'Monthly, and at every data drop'),
]
r = 10
for k, v in meta:
    ws.cell(row=r, column=2, value=k).font = lbl_f
    ws.cell(row=r, column=3, value=v).font = body_f
    r += 1

r += 1
ws.cell(row=r, column=2, value='How to use this workbook').font = h_f
r += 1
for line in [
    ('MIDP', 'The project register. One row per deliverable. Filter by discipline, stage or status. '
             'The Information Manager maintains this sheet.'),
    ('TIDP', 'The return template. Each appointed party completes a copy for its own scope and returns it '
             'to the Information Manager, who merges it into the MIDP sheet.'),
    ('Summary', 'Counts by stage and by discipline, calculated from the MIDP sheet. Nothing is typed here.'),
    ('Lists', 'The permitted values behind the drop-down lists. Amend here to change the drop-downs.'),
]:
    ws.cell(row=r, column=2, value=line[0]).font = lbl_f
    c = ws.cell(row=r, column=3, value=line[1])
    c.font = body_f
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 28
    r += 1

r += 1
ws.cell(row=r, column=2, value='Before this plan is baselined').font = h_f
r += 1
note = ('The Originator column is deliberately empty. The originator code register has not been confirmed, '
        'and the container naming rule requires a code of a fixed length that is still to be agreed. Complete '
        'the Originator column only once the register is issued. Nothing on this project is to be numbered '
        'before that point.')
c = ws.cell(row=r, column=2, value=note)
c.font = Font(name='Calibri', size=9.5)
c.alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=3)
for rr in range(r, r + 3):
    for cc in (2, 3):
        ws.cell(row=rr, column=cc).fill = shade_fill
r += 4

ws.cell(row=r, column=2, value=('Issued through the Common Data Environment. Uncontrolled when printed.')).font = small_f

# ═══ 2. MIDP register ═══════════════════════════════════════════════════════
ms = wb.create_sheet('MIDP')
ms.sheet_view.showGridLines = False
for i, (name, width) in enumerate(COLS, start=1):
    ms.cell(row=1, column=i, value=name)
    ms.column_dimensions[get_column_letter(i)].width = width
style_header(ms, 1, len(COLS))

for n, row in enumerate(R, start=2):
    (ref, disc, deliv, typ, stage, lod, fmt, suit, state, month, resp, tidp, notes) = row
    vals = [ref, disc, '', deliv, typ, stage, lod, fmt, suit, state, month, None, None,
            '=IF(AND(L{0}<>"",M{0}<>""),M{0}-L{0},"")'.format(n), resp, tidp, '', notes]
    for i, v in enumerate(vals, start=1):
        c = ms.cell(row=n, column=i, value=v)
        c.font = body_f
        c.border = box
        c.alignment = Alignment(vertical='top', wrap_text=(i in (4, 18)))
        if i in (12, 13):
            c.number_format = 'dd mmm yyyy'
        if i == 14:
            c.number_format = '0;-0;""'
            c.alignment = Alignment(horizontal='center', vertical='top')
        if i in (7, 9, 10, 17):
            c.alignment = Alignment(horizontal='center', vertical='top')

last = len(R) + 1
ms.freeze_panes = 'A2'
ms.auto_filter.ref = 'A1:R%d' % last

# Drop-downs. The source column on the Lists sheet is NOT the same letter as the
# column being validated -- LIST_COL maps them explicitly, in midp_schema.py.
# Pointing a list at its own column letter silently offers the wrong values,
# which a reader would take as correct.
for col_letter, key in MIDP_VALIDATED:
    src = LIST_COL[key]
    dv = DataValidation(type='list', formula1="'Lists'!$%s$2:$%s$30" % (src, src),
                        allow_blank=True, showDropDown=False)
    dv.error = 'Select a value from the list. To add a permitted value, amend the Lists sheet.'
    dv.errorTitle = 'Value not permitted'
    ms.add_data_validation(dv)
    dv.add('%s2:%s%d' % (col_letter, col_letter, last))

# RAG colouring
rag = 'Q2:Q%d' % last
ms.conditional_formatting.add(rag, CellIsRule(operator='equal', formula=['"Red"'], fill=PatternFill('solid', fgColor=RED)))
ms.conditional_formatting.add(rag, CellIsRule(operator='equal', formula=['"Amber"'], fill=PatternFill('solid', fgColor=AMBER)))
ms.conditional_formatting.add(rag, CellIsRule(operator='equal', formula=['"Green"'], fill=PatternFill('solid', fgColor=GREEN)))
ms.conditional_formatting.add(rag, CellIsRule(operator='equal', formula=['"Complete"'], fill=PatternFill('solid', fgColor=GREEN)))
# overdue: planned date passed and no actual date
ms.conditional_formatting.add(
    'A2:R%d' % last,
    FormulaRule(formula=['AND($L2<>"",$M2="",$L2<TODAY())'], fill=PatternFill('solid', fgColor=RED), stopIfTrue=False))

ms.page_setup.orientation = 'landscape'
ms.page_setup.fitToWidth = 1
ms.page_setup.fitToHeight = 0
ms.sheet_properties.pageSetUpPr.fitToPage = True
ms.print_title_rows = '1:1'

# ═══ 3. TIDP return template ════════════════════════════════════════════════
ts = wb.create_sheet('TIDP')
ts.sheet_view.showGridLines = False
ts.column_dimensions['A'].width = 3
ts['B2'] = 'Task Information Delivery Plan'
ts['B2'].font = Font(name='Calibri', size=16, bold=True, color=NAVY)
ts['B4'] = ('Complete one copy of this sheet for your scope and return it to the Information Manager. '
            'Rows are merged into the project Master Information Delivery Plan. Use the drop-down lists; '
            'if a value you need is not offered, raise it rather than typing a variant.')
ts['B4'].font = Font(name='Calibri', size=9.5)
ts['B4'].alignment = Alignment(wrap_text=True, vertical='top')
ts.merge_cells('B4:H5')

hdr = [('Appointed party', '[FILL]'), ('Discipline', '[FILL]'), ('Originator code', '[FILL — once the register is issued]'),
       ('Task Team Manager', '[FILL]'), ('Contact', '[FILL]'), ('TIDP reference', '[FILL — e.g. TIDP-A]'),
       ('Revision', 'P01'), ('Date', '[FILL]')]
r = 7
for k, v in hdr:
    ts.cell(row=r, column=2, value=k).font = lbl_f
    ts.cell(row=r, column=3, value=v).font = body_f
    r += 1

r += 1
for i, (name, width) in enumerate(COLS, start=2):
    ts.cell(row=r, column=i, value=name)
    ts.column_dimensions[get_column_letter(i)].width = width
style_header(ts, r, len(COLS) + 1)
blank_start = r + 1
for n in range(blank_start, blank_start + 40):
    for i in range(2, len(COLS) + 2):
        c = ts.cell(row=n, column=i)
        c.border = box
        c.font = body_f
        if i in (13, 14):
            c.number_format = 'dd mmm yyyy'
for col_letter, key in TIDP_VALIDATED:
    src = LIST_COL[key]
    dv = DataValidation(type='list', formula1="'Lists'!$%s$2:$%s$30" % (src, src), allow_blank=True,
                        showDropDown=False)
    ts.add_data_validation(dv)
    dv.add('%s%d:%s%d' % (col_letter, blank_start, col_letter, blank_start + 39))
ts.freeze_panes = 'A%d' % blank_start
ts.page_setup.orientation = 'landscape'
ts.page_setup.fitToWidth = 1
ts.sheet_properties.pageSetUpPr.fitToPage = True

# ═══ 4. Summary ═════════════════════════════════════════════════════════════
ss = wb.create_sheet('Summary')
ss.sheet_view.showGridLines = False
ss.column_dimensions['A'].width = 3
ss.column_dimensions['B'].width = 26
for col in 'CDEFG':
    ss.column_dimensions[col].width = 14

ss['B2'] = 'Delivery summary'
ss['B2'].font = Font(name='Calibri', size=16, bold=True, color=NAVY)
ss['B3'] = 'Calculated from the MIDP sheet. Do not type in this sheet.'
ss['B3'].font = small_f

ss['B5'] = 'By stage'
ss['B5'].font = h_f
heads = ['Stage', 'Deliverables', 'Complete', 'Red', 'Amber', 'Outstanding']
for i, hh in enumerate(heads, start=2):
    ss.cell(row=6, column=i, value=hh)
style_header(ss, 6, 7)
r = 7
for stage in LISTS['Stage']:
    ss.cell(row=r, column=2, value=stage).font = body_f
    ss.cell(row=r, column=3, value='=COUNTIF(MIDP!$F:$F,$B%d)' % r).font = body_f
    ss.cell(row=r, column=4, value='=COUNTIFS(MIDP!$F:$F,$B%d,MIDP!$Q:$Q,"Complete")' % r).font = body_f
    ss.cell(row=r, column=5, value='=COUNTIFS(MIDP!$F:$F,$B%d,MIDP!$Q:$Q,"Red")' % r).font = body_f
    ss.cell(row=r, column=6, value='=COUNTIFS(MIDP!$F:$F,$B%d,MIDP!$Q:$Q,"Amber")' % r).font = body_f
    ss.cell(row=r, column=7, value='=C%d-D%d' % (r, r)).font = body_f
    for i in range(2, 8):
        ss.cell(row=r, column=i).border = box
        if i > 2:
            ss.cell(row=r, column=i).alignment = Alignment(horizontal='center')
    r += 1
ss.cell(row=r, column=2, value='Total').font = lbl_f
for i, col in enumerate('CDEFG', start=3):
    ss.cell(row=r, column=i, value='=SUM(%s7:%s%d)' % (col, col, r - 1)).font = lbl_f
    ss.cell(row=r, column=i).alignment = Alignment(horizontal='center')
for i in range(2, 8):
    ss.cell(row=r, column=i).border = box
    ss.cell(row=r, column=i).fill = shade_fill

r += 3
ss.cell(row=r, column=2, value='By discipline').font = h_f
r += 1
for i, hh in enumerate(['Discipline', 'Deliverables', 'Complete', 'Red', 'Amber', 'Outstanding'], start=2):
    ss.cell(row=r, column=i, value=hh)
style_header(ss, r, 7)
r += 1
for disc in LISTS['Discipline']:
    ss.cell(row=r, column=2, value=disc).font = body_f
    ss.cell(row=r, column=3, value='=COUNTIF(MIDP!$B:$B,$B%d)' % r).font = body_f
    ss.cell(row=r, column=4, value='=COUNTIFS(MIDP!$B:$B,$B%d,MIDP!$Q:$Q,"Complete")' % r).font = body_f
    ss.cell(row=r, column=5, value='=COUNTIFS(MIDP!$B:$B,$B%d,MIDP!$Q:$Q,"Red")' % r).font = body_f
    ss.cell(row=r, column=6, value='=COUNTIFS(MIDP!$B:$B,$B%d,MIDP!$Q:$Q,"Amber")' % r).font = body_f
    ss.cell(row=r, column=7, value='=C%d-D%d' % (r, r)).font = body_f
    for i in range(2, 8):
        ss.cell(row=r, column=i).border = box
        if i > 2:
            ss.cell(row=r, column=i).alignment = Alignment(horizontal='center')
    r += 1

r += 2
ss.cell(row=r, column=2, value='Overdue (planned date passed, no actual date)').font = lbl_f
ss.cell(row=r, column=3,
        value='=SUMPRODUCT((MIDP!$L$2:$L$%d<>"")*(MIDP!$M$2:$M$%d="")*(MIDP!$L$2:$L$%d<TODAY()))'
              % (last, last, last)).font = Font(name='Calibri', size=9.5, bold=True, color='B00020')
ss.cell(row=r, column=3).alignment = Alignment(horizontal='center')

# ═══ 5. Lists ═══════════════════════════════════════════════════════════════
ls = wb.create_sheet('Lists')
ls.sheet_view.showGridLines = False
ls['A1'] = 'Permitted values. Amend here to change the drop-down lists on the MIDP and TIDP sheets.'
ls['A1'].font = small_f
order = ['Discipline', 'Originator', 'Type', 'Stage', 'LOD', 'Suitability', 'CDE State', 'RAG']
for i, key in enumerate(order, start=1):
    col = get_column_letter(i)
    ls.column_dimensions[col].width = 22
    c = ls.cell(row=1, column=i, value=key if key in LISTS or key == 'Originator' else key)
    c.font = Font(name='Calibri', size=9, bold=True, color=NAVY)
    c.fill = hdr_fill
    c.border = box
    vals = LISTS.get(key, ['[FILL — from the originator code register]'])
    for j, v in enumerate(vals, start=2):
        cc = ls.cell(row=j, column=i, value=v)
        cc.font = body_f
        cc.border = box
ls['A1'] = 'Discipline'
ls['A1'].font = Font(name='Calibri', size=9, bold=True, color=NAVY)
ls['A1'].fill = hdr_fill
ls['A1'].border = box

wb.properties.title = 'KUT Master Information Delivery Plan'
wb.properties.subject = 'Kampala Uganda Temple — aggregated information delivery schedule'
wb.properties.creator = 'Symbion Consulting Group Studios'
wb.properties.lastModifiedBy = 'Symbion Consulting Group Studios'
wb.properties.category = 'Project procedure'

# Determinism. openpyxl stamps dcterms:created and dcterms:modified from the wall
# clock, so an unpinned build produced a different binary every run -- and that
# was true of the committed workbook until this was added. An always-dirty binary
# diff trains reviewers to ignore `git status` on exactly the file a hand-edit
# would show up in. Pinned, regeneration is a genuine no-op.
_EPOCH = datetime.datetime(*K.EPOCH)
wb.properties.created = _EPOCH
wb.properties.modified = _EPOCH

# The staleness stamp rides in dc:description, which openpyxl already writes, so
# the gate can read it back with plain `zipfile` and stay stdlib-only.
_ROOT = pathlib.Path(__file__).resolve().parent.parent
wb.properties.description = K.with_provenance(
    'Rev P01. Issued through the Common Data Environment. Uncontrolled when printed.',
    'tools/build_midp.py', K.inputs_digest(_ROOT, OUT))

wb.active = 0
wb.save(OUT)
K.finalise(pathlib.Path(OUT))
print('saved:', OUT, '|', len(R), 'deliverables')
