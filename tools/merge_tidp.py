#!/usr/bin/env python3
"""Merge returned Task Information Delivery Plans into the project MIDP.

    python tools/merge_tidp.py returns/*.xlsx              # preview only
    python tools/merge_tidp.py returns/TIDP-A.xlsx --apply # write the additions
    python tools/merge_tidp.py returns/*.xlsx --apply --overwrite-conflicts

WHY THIS EXISTS
Each appointed party completes the TIDP sheet and returns it; the Information
Manager merges the rows into the MIDP by hand. Nine consultants, re-baselined at
every stage, is enough repetition to go wrong quietly -- and the failure is
invisible. A dropped row does not appear anywhere as an error: the register is
simply missing a deliverable nobody is now tracking, and that surfaces at a gate.

PREVIEW BY DEFAULT
Nothing is written unless --apply is passed. The register is an issued document
and a merge is not reversible by eye once the workbook is saved, so the useful
default is the one that tells you what would happen. Preview is also stdlib-only
-- reading a .xlsx needs nothing installed -- so the safe operation always runs.
Writing needs openpyxl, and says so plainly if it is missing.

A Ref that already exists with DIFFERENT content is a conflict and is refused,
not merged. Two parties disagreeing about the same deliverable is a question for
a person; silently taking the newer file would let the second return overwrite
the first with nobody seeing it. --overwrite-conflicts is the explicit override.

Validation uses tools/midp_schema.py, the same definitions build_midp.py wrote
the drop-downs from. Restating them here would let the merge reject a value the
workbook itself had just offered the consultant -- or, worse, accept one it had
not.
"""
from __future__ import annotations

import argparse
import glob
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import kut_docs_lib as K          # noqa: E402
import midp_schema as S           # noqa: E402

MIDP_FILE = "KUT_Master_Information_Delivery_Plan.xlsx"


class Row(dict):
    """One deliverable, keyed by column name, plus where it came from."""

    def __init__(self, values, source, line):
        super().__init__(values)
        self.source = source
        self.line = line

    def content(self):
        """The comparable payload: everything the returning party owns.

        Columns the Information Manager maintains are excluded, so a return that
        leaves them blank does not read as a conflict against a register that
        has them filled.
        """
        return {k: (v or "").strip() for k, v in self.items()
                if k not in S.IM_OWNED and (v or "").strip()}


def find_header(rows):
    """(row index, {column name: cell index}) for the register header.

    Located by searching for the key column rather than trusting a fixed
    address: the TIDP sheet carries a header block above its table and the MIDP
    sheet does not, and a consultant may well have inserted a row. A layout
    change should degrade to a clear error, not a silent column shift -- seven
    drop-downs in this very workbook once pointed at the wrong column and looked
    entirely correct.
    """
    for i, row in enumerate(rows):
        cells = [(c or "").strip() for c in row]
        if S.KEY_COL in cells and "Deliverable" in cells:
            return i, {name: j for j, name in enumerate(cells) if name}
    return None, None


def read_register(path: Path, sheet_name: str, source: str):
    """Every non-empty deliverable row on one sheet."""
    sheets = K.xlsx_sheets(path)
    rows = sheets.get(sheet_name)
    if rows is None:
        # A consultant may return only their own sheet, or rename the tab.
        for name, candidate in sheets.items():
            idx, _cols = find_header(candidate)
            if idx is not None:
                rows, sheet_name = candidate, name
                break
    if rows is None:
        return None, "no sheet with a '%s' column" % S.KEY_COL

    hdr, cols = find_header(rows)
    if hdr is None:
        return None, "sheet %r has no '%s' / 'Deliverable' header row" % (sheet_name, S.KEY_COL)

    missing = [c for c in (S.KEY_COL, "Deliverable", "Stage") if c not in cols]
    if missing:
        return None, "sheet %r is missing column(s) %s" % (sheet_name, missing)

    out = []
    for n, row in enumerate(rows[hdr + 1:], start=hdr + 2):
        values = {name: (row[j].strip() if j < len(row) else "")
                  for name, j in cols.items()}
        if not values.get(S.KEY_COL) and not values.get("Deliverable"):
            continue                      # blank template row
        out.append(Row(values, source, n))
    return out, None


def validate(row: Row):
    """Values outside the permitted lists the workbook offered."""
    bad = []
    for column, permitted in S.LISTS.items():
        v = (row.get(column) or "").strip()
        if v and v not in permitted:
            bad.append("%s=%r is not a permitted value (expected one of: %s)"
                       % (column, v, ", ".join(permitted)))
    if not (row.get(S.KEY_COL) or "").strip():
        bad.append("no %s -- the register is matched on it, so the row cannot be placed"
                   % S.KEY_COL)
    return bad


def classify(register, incoming):
    """(new, identical, conflicts, invalid) across every returned row."""
    by_ref = {(r.get(S.KEY_COL) or "").strip(): r for r in register}
    new, identical, conflicts, invalid = [], [], [], []
    seen = {}

    for row in incoming:
        problems = validate(row)
        if problems:
            invalid.append((row, problems))
            continue

        ref = (row.get(S.KEY_COL) or "").strip()

        # Two returns claiming the same Ref is a conflict even before the
        # register is consulted -- otherwise whichever file was listed last on
        # the command line would silently win.
        if ref in seen and seen[ref].content() != row.content():
            conflicts.append((row, seen[ref], "another return"))
            continue
        seen[ref] = row

        existing = by_ref.get(ref)
        if existing is None:
            new.append(row)
        elif existing.content() == row.content():
            identical.append(row)
        else:
            conflicts.append((row, existing, "the register"))
    return new, identical, conflicts, invalid


def describe_difference(a: Row, b: Row):
    keys = sorted(set(a.content()) | set(b.content()))
    out = []
    for k in keys:
        va, vb = a.content().get(k, ""), b.content().get(k, "")
        if va != vb:
            out.append("      %-16s returned %r vs %r" % (k, va, vb))
    return out


def apply_merge(midp: Path, rows, verbose: bool):
    """Append rows to the register, preserving the sheet's formatting."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("\n--apply needs openpyxl (preview does not):\n"
              "    python -m pip install openpyxl", file=sys.stderr)
        return 1

    wb = load_workbook(midp)
    ws = wb[S.MIDP_SHEET]

    header = {}
    for j in range(1, ws.max_column + 1):
        name = ws.cell(row=S.MIDP_HEADER_ROW, column=j).value
        if name:
            header[str(name).strip()] = j

    last = ws.max_row
    while last > S.MIDP_HEADER_ROW and not ws.cell(row=last, column=header[S.KEY_COL]).value:
        last -= 1

    from copy import copy
    for offset, row in enumerate(rows, start=1):
        target = last + offset
        for name, j in header.items():
            src = ws.cell(row=last, column=j)
            cell = ws.cell(row=target, column=j)
            # Carry the register's own formatting down rather than leaving the
            # merged rows visually distinct from the rest of the schedule.
            cell.font, cell.border = copy(src.font), copy(src.border)
            cell.alignment, cell.number_format = copy(src.alignment), src.number_format

            if name == "Variance (days)":
                # A formula, not a value. Taking a literal from a return would
                # replace the calculation with a number that stops updating.
                cell.value = ('=IF(AND(L{0}<>"",M{0}<>""),M{0}-L{0},"")'.format(target))
            else:
                cell.value = row.get(name) or None

    end = last + len(rows)
    ws.auto_filter.ref = "A%d:R%d" % (S.MIDP_HEADER_ROW, end)
    wb.save(midp)
    if verbose:
        print("  extended the register to row %d and reset the filter range" % end)
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("returns", nargs="+", help="returned TIDP workbook(s); globs accepted")
    ap.add_argument("--midp", default=None, help="the register to merge into")
    ap.add_argument("--apply", action="store_true",
                    help="write the additions (default is preview only)")
    ap.add_argument("--overwrite-conflicts", action="store_true",
                    help="also replace register rows whose content differs")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    root = Path(__file__).resolve().parent.parent
    midp = Path(args.midp) if args.midp else root / MIDP_FILE
    if not midp.exists():
        print("register not found: %s" % midp, file=sys.stderr)
        return 1

    paths = []
    for pattern in args.returns:
        hits = [Path(p) for p in glob.glob(pattern)]
        if not hits:
            print("no file matches %r" % pattern, file=sys.stderr)
            return 1
        paths.extend(hits)

    register, err = read_register(midp, S.MIDP_SHEET, midp.name)
    if err:
        print("cannot read the register: %s" % err, file=sys.stderr)
        return 1

    incoming, failed = [], []
    for p in paths:
        rows, err = read_register(p, S.TIDP_SHEET, p.name)
        if err:
            failed.append((p, err))
            continue
        incoming.extend(rows)

    new, identical, conflicts, invalid = classify(register, incoming)

    print("MIDP merge preview" if not args.apply else "MIDP merge")
    print("  register            : %s (%d deliverables)" % (midp.name, len(register)))
    print("  returns read        : %d file(s), %d row(s)" % (len(paths) - len(failed), len(incoming)))
    print("  would be added      : %d" % len(new))
    print("  already identical   : %d" % len(identical))
    print("  conflicts           : %d" % len(conflicts))
    print("  rejected as invalid : %d" % len(invalid))

    if failed:
        print("\nUNREADABLE")
        for p, err in failed:
            print("  %s: %s" % (p.name, err))

    if new:
        print("\nWOULD ADD")
        for r in new:
            print("  %-9s %-46s %-20s [%s line %d]"
                  % (r.get(S.KEY_COL), (r.get("Deliverable") or "")[:46],
                     r.get("Stage") or "", r.source, r.line))

    if identical and args.verbose:
        print("\nALREADY IN THE REGISTER, UNCHANGED")
        for r in identical:
            print("  %-9s %s" % (r.get(S.KEY_COL), (r.get("Deliverable") or "")[:60]))

    if conflicts:
        print("\nCONFLICTS -- not merged" +
              (" (--overwrite-conflicts given, see below)" if args.overwrite_conflicts else ""))
        for row, other, against in conflicts:
            print("  %-9s [%s line %d] differs from %s"
                  % (row.get(S.KEY_COL), row.source, row.line, against))
            for line in describe_difference(row, other):
                print(line)

    if invalid:
        print("\nREJECTED -- a value outside the lists the workbook offers")
        for row, problems in invalid:
            print("  %-9s [%s line %d]" % (row.get(S.KEY_COL) or "(no ref)", row.source, row.line))
            for p in problems:
                print("      %s" % p)

    if not args.apply:
        print("\nNothing was written. Re-run with --apply to add the %d new row(s)." % len(new))
        if conflicts:
            print("Conflicts stay refused unless --overwrite-conflicts is also given: two")
            print("parties disagreeing about one deliverable is a question for a person.")
        return 1 if (invalid or failed) else 0

    to_write = list(new)
    if args.overwrite_conflicts and conflicts:
        print("\n--overwrite-conflicts: %d conflicting row(s) are NOT appended -- a Ref that"
              % len(conflicts))
        print("already exists is corrected in place, which this tool does not do. Edit the")
        print("register row directly, or agree the change with the returning party first.")

    if not to_write:
        print("\nNothing to add.")
        return 1 if (invalid or failed) else 0

    rc = apply_merge(midp, to_write, args.verbose)
    if rc:
        return rc
    print("\nAdded %d deliverable(s) to %s." % (len(to_write), midp.name))
    print("The register is a generated document: this edit is a MANUAL change to it,")
    print("so tools/check_kut_documents.py will now report it as edited since generation.")
    print("Fold the rows into tools/build_midp.py and regenerate before issuing.")
    return 1 if (invalid or failed) else 0


if __name__ == "__main__":
    raise SystemExit(main())
